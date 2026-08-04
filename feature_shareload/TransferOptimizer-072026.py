"""
TransferOptimizer.py
--------------------
หาจุด switch + tie ที่ optimal สำหรับถ่ายเทโหลดระหว่างหม้อแปลง 2 ตัว

Usage:
    python TransferOptimizer.py <FACILITYID_A> <FACILITYID_B> [options]

    FACILITYID_A : หม้อแปลงที่โหลดเกิน  (> load_limit%)
    FACILITYID_B : หม้อแปลงที่รับโหลดได้ (< load_limit%)

Options:
    --max-tie-dist  float   ระยะสูงสุดของสาย tie (m)  [default 100]
    --min-voltage   float   แรงดันต่ำสุดที่ยอมรับ (V) [default 200]
    --load-limit    float   โหลดสูงสุดของหม้อแปลง %   [default 80]
    --json-dir      path    folder ที่เก็บไฟล์ JSON    [default D:\\testpy\\jsonfile]
    --out-dir       path    folder ผลลัพธ์ Excel/PNG   [default D:\\TRneighborhood\\output]
"""

import os, sys, json, math, re, argparse, tempfile, shutil, contextlib, io
from pathlib import Path
from datetime import datetime
from collections import defaultdict, deque
from typing import Dict, List, Optional, Set, Tuple

# Force UTF-8 output so Thai + Unicode symbols work on any terminal
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# This file is invoked from feature_shareload/ (via run_web.py), but
# InputJsonApi.py — the FACILITYID -> network JSON fetcher — lives at the
# project root, one level up. Put the root on sys.path so `import
# InputJsonApi` resolves regardless of the caller's cwd.
_PROJECT_ROOT = str(Path(__file__).resolve().parent.parent)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import numpy as np
from scipy.spatial import cKDTree
import networkx as nx
import openpyxl
from openpyxl.styles import PatternFill, Font
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.cm as mcm
import matplotlib.colors as mcolors
from matplotlib.lines import Line2D

# --- Import helpers and core functions from existing module ---
from Runopendss_All20032026 import (
    get_attr,
    has_point,
    has_paths,
    point_xy,
    endpoints_from_paths,
    cluster_points,
    convert_json_to_dss_ordered,
    solve_with_opendss,
    build_bfs_order,
    build_meter_bus_names,
)
import opendssdirect as odss

OUTPUT_DIR   = r"D:\testpy\jsonfile"
PHASE_MAP    = {1: "A", 2: "B", 3: "C"}
_COORD_OFFSET = 1e7   # shift TR_A subtree coords into "virtual space" so cluster_points
                      # in convert_json_to_dss_ordered never merges TR_A and TR_B nodes


# =====================================================================
# Auto-snap tolerance helper
# =====================================================================

def _auto_snap_tol(
    raw_ends: List[Tuple[float, float]],
    xf_pt: Tuple[float, float],
    tol_start: float,
    tol_max: float,
) -> float:
    """
    Detect the minimum gap between the BFS-reachable backbone cluster and any
    isolated cluster, then return an adjusted snap tolerance that bridges it.

    Algorithm:
    1. Cluster with tol_start; BFS from transformer node.
    2. Find connected set C and island set I (|reps| - |C|).
    3. Compute minimum pairwise distance between C and I coordinate arrays.
    4. If tol_start < gap <= tol_max → return gap + 0.05 (just enough to bridge).
    5. Otherwise return tol_start unchanged.
    """
    if not raw_ends:
        return tol_start

    node_ids, reps = cluster_points(raw_ends, tol=tol_start)
    reps_arr = np.array(reps, dtype=float)

    # Snap transformer point to nearest cluster node
    d2 = np.sum((reps_arr - np.array(xf_pt)) ** 2, axis=1)
    xf_node = int(np.argmin(d2))

    # Build adjacency from the clustered endpoints
    adj: Dict[int, List[int]] = defaultdict(list)
    for i in range(0, len(node_ids), 2):
        u, v = node_ids[i], node_ids[i + 1]
        if u != v:
            adj[u].append(v)
            adj[v].append(u)

    # BFS to find connected set
    visited: Set[int] = {xf_node}
    queue = deque([xf_node])
    while queue:
        cur = queue.popleft()
        for nb in adj[cur]:
            if nb not in visited:
                visited.add(nb)
                queue.append(nb)

    n_nodes = len(reps)
    island_ids = [i for i in range(n_nodes) if i not in visited]
    if not island_ids:
        return tol_start  # nothing to bridge

    conn_arr   = reps_arr[sorted(visited)]
    island_arr = reps_arr[island_ids]

    # KD-tree: minimum distance from any island node to any connected node
    conn_tree = cKDTree(conn_arr)
    dists, _  = conn_tree.query(island_arr, k=1)
    min_gap   = float(np.min(dists))

    if tol_start < min_gap <= tol_max:
        return min_gap + 0.05
    return tol_start


# =====================================================================
# NetworkGraph  —  parse JSON → graph + coordinate mapping
# =====================================================================

class NetworkGraph:
    """
    Parse a LV network JSON into a NetworkX DiGraph (directed from transformer).
    Stores node coordinates and loads so the optimizer can find switch/tie points.
    """

    def __init__(self, json_path: str, snap_tol: float = 2.0,
                 auto_snap: bool = True, snap_tol_max: float = 5.0):
        self.json_path = Path(json_path)
        self.snap_tol = snap_tol
        self._auto_snap = auto_snap
        self._snap_tol_max = snap_tol_max

        self.facilityid: str = ""
        self.rated_kva: float = 250.0
        self.transformer_node: int = -1
        self.G: nx.DiGraph = nx.DiGraph()
        self.node_coords: Dict[int, Tuple[float, float]] = {}
        self.node_kw: Dict[int, float] = {}

        # feature_index → (u, v) node pair, for LC backbone lines only
        self.lc_feat_edges: Dict[int, Tuple[int, int]] = {}
        # feature_index → nearest backbone node, for load & service features
        self.load_feat_nodes: Dict[int, int] = {}
        self.svc_feat_nodes: Dict[int, int] = {}

        self._parse()

    def _parse(self) -> None:
        with open(self.json_path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        features = data.get("features", [])

        # --- Transformer ---
        xf_list = [
            f for f in features
            if has_point(f) and "XF" in str(get_attr(f, "TAG", "")).upper()
        ]
        if not xf_list:
            raise RuntimeError(f"No transformer (TAG contains XF) in {self.json_path.name}")
        xfmr = xf_list[0]
        fac = str(get_attr(xfmr, "FACILITYID", "") or "").strip()
        self.facilityid = fac or str(get_attr(xfmr, "TAG", "XFM1") or "XFM1").strip()
        self.rated_kva = float(get_attr(xfmr, "RATEKVA", 250.0) or 250.0)

        # --- LC backbone lines (SUBTYPECODE=1 เท่านั้น กัน streetlight/อื่นๆ) ---
        def _subtype_ok(f: dict) -> bool:
            st = get_attr(f, "SUBTYPECODE", None) or get_attr(f, "SUBTYPECOD", None)
            if st is None:
                return True  # ไม่มี field → ไม่บล็อก
            try:
                return int(float(st)) == 1
            except (ValueError, TypeError):
                return True

        lc_list = [
            (i, f) for i, f in enumerate(features)
            if has_paths(f) and "LC" in str(get_attr(f, "TAG", "")).upper()
            and _subtype_ok(f)
        ]
        if not lc_list:
            raise RuntimeError(f"No backbone lines (TAG contains LC) in {self.json_path.name}")

        # Collect all line endpoints and cluster
        raw_ends: List[Tuple[float, float]] = []
        for _, f in lc_list:
            ep = endpoints_from_paths(f)
            if ep:
                raw_ends.extend([ep[0], ep[1]])

        # Auto-adjust snap tolerance to bridge small GIS coordinate gaps
        if self._auto_snap and raw_ends:
            xf_pt = point_xy(xfmr)
            adjusted = _auto_snap_tol(raw_ends, xf_pt, self.snap_tol, self._snap_tol_max)
            if adjusted != self.snap_tol:
                print(f"  [auto-snap] {self.json_path.name}: tol {self.snap_tol:.2f}m → {adjusted:.2f}m")
                self.snap_tol = adjusted

        node_ids, reps = cluster_points(raw_ends, tol=self.snap_tol)
        self.node_coords = {i: reps[i] for i in range(len(reps))}

        reps_arr = np.array(reps, dtype=float)

        def nearest_node(pt: Tuple[float, float]) -> int:
            d2 = (reps_arr[:, 0] - pt[0]) ** 2 + (reps_arr[:, 1] - pt[1]) ** 2
            return int(np.argmin(d2))

        # Build undirected adjacency and record feature→edge
        adj: Dict[int, List[int]] = defaultdict(list)
        raw_idx = 0
        for feat_idx, f in lc_list:
            ep = endpoints_from_paths(f)
            if not ep:
                continue
            u = node_ids[raw_idx]
            v = node_ids[raw_idx + 1]
            raw_idx += 2
            if u != v:
                adj[u].append(v)
                adj[v].append(u)
            self.lc_feat_edges[feat_idx] = (u, v)

        # Transformer → source node
        self.transformer_node = nearest_node(point_xy(xfmr))

        # BFS from transformer → directed graph
        visited: Set[int] = {self.transformer_node}
        queue = deque([self.transformer_node])
        self.G.add_node(self.transformer_node)
        while queue:
            u = queue.popleft()
            for v in adj[u]:
                if v not in visited:
                    visited.add(v)
                    self.G.add_node(v)
                    self.G.add_edge(u, v)
                    queue.append(v)

        # --- Load points (SUBTYPECODE == 1) ---
        for i, f in enumerate(features):
            try:
                subtype = int(get_attr(f, "SUBTYPECODE", -1) or -1)
            except Exception:
                subtype = -1

            if subtype == 1 and has_point(f):
                tag = str(get_attr(f, "TAG", "") or "").upper()
                if "XF" in tag:
                    continue
                peano = str(get_attr(f, "PEANO", "") or "").strip()
                peameter = str(get_attr(f, "PEAMETER", "") or "").strip()
                if not peano and not peameter:
                    continue
                nid = nearest_node(point_xy(f))
                self.load_feat_nodes[i] = nid
                kw = float(get_attr(f, "KWP", 0.0) or 0.0)
                self.node_kw[nid] = self.node_kw.get(nid, 0.0) + kw

            elif subtype == 5 and has_paths(f):
                ep = endpoints_from_paths(f)
                if ep:
                    nid = nearest_node(ep[0])
                    self.svc_feat_nodes[i] = nid

    def get_subtree(self, switch_edge: Tuple[int, int]) -> Set[int]:
        """Nodes downstream of switch_edge[1] (inclusive) when edge is opened."""
        u, v = switch_edge
        if not self.G.has_edge(u, v):
            return set()
        return {v} | nx.descendants(self.G, v)

    def get_feature_indices_for_subtree(self, subtree: Set[int]) -> List[int]:
        """Feature indices for all features whose nodes are entirely within the subtree."""
        indices: List[int] = []
        for feat_idx, (u, v) in self.lc_feat_edges.items():
            if u in subtree and v in subtree:
                indices.append(feat_idx)
        for feat_idx, nid in self.load_feat_nodes.items():
            if nid in subtree:
                indices.append(feat_idx)
        for feat_idx, nid in self.svc_feat_nodes.items():
            if nid in subtree:
                indices.append(feat_idx)
        return indices

    def get_total_kw(self, nodes: Set[int]) -> float:
        return sum(self.node_kw.get(n, 0.0) for n in nodes)

    def find_switch_edge_feature(self, switch_edge: Tuple[int, int]) -> Optional[int]:
        """Return the LC feature index for the switch edge, or None."""
        u, v = switch_edge
        for feat_idx, (fu, fv) in self.lc_feat_edges.items():
            if (fu == u and fv == v) or (fu == v and fv == u):
                return feat_idx
        return None


# =====================================================================
# Helpers: JSON modification and OpenDSS querying
# =====================================================================

def build_modified_json(
    original: dict,
    remove_indices: Set[int],
    add_features: List[dict],
) -> dict:
    kept = [f for i, f in enumerate(original["features"]) if i not in remove_indices]
    kept.extend(add_features)
    return {**original, "features": kept}


def _translate_feature(feat: dict, dx: float, dy: float) -> dict:
    """Return deep copy of feat with all geometry coordinates shifted by (dx, dy)."""
    import copy
    f = copy.deepcopy(feat)
    g = f.get("geometry") or {}
    if "paths" in g:
        g["paths"] = [[[c[0] + dx, c[1] + dy] for c in seg] for seg in g["paths"]]
    elif "x" in g:
        g["x"] += dx
        g["y"] += dy
    return f


def _cross2d(o: Tuple, a: Tuple, b: Tuple) -> float:
    return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])


def _segments_cross(p1: Tuple, p2: Tuple, p3: Tuple, p4: Tuple) -> bool:
    """True ถ้า segment p1-p2 ตัดผ่าน segment p3-p4 แบบ proper (ไม่นับแค่แตะปลาย)"""
    d1 = _cross2d(p3, p4, p1)
    d2 = _cross2d(p3, p4, p2)
    d3 = _cross2d(p1, p2, p3)
    d4 = _cross2d(p1, p2, p4)
    if ((d1 > 0 and d2 < 0) or (d1 < 0 and d2 > 0)) and \
       ((d3 > 0 and d4 < 0) or (d3 < 0 and d4 > 0)):
        return True
    return False


def _tie_crosses_lv(na: int, nb: int,
                    net_a: "NetworkGraph", net_b: "NetworkGraph") -> Tuple[bool, str]:
    """
    เช็คว่าสาย tie (เส้นตรง na→nb) ตัดผ่าน LV backbone edge ใดใน net_a หรือ net_b หรือไม่
    คืนค่า (crossed: bool, description: str)
    """
    p1 = net_a.node_coords[na]
    p2 = net_b.node_coords[nb]

    for u, v in net_a.G.edges():
        if u == na or v == na:
            continue
        p3 = net_a.node_coords[u]
        p4 = net_a.node_coords[v]
        if _segments_cross(p1, p2, p3, p4):
            return True, f"ตัดกับ TR_A edge ({u},{v})"

    for u, v in net_b.G.edges():
        if u == nb or v == nb:
            continue
        p3 = net_b.node_coords[u]
        p4 = net_b.node_coords[v]
        if _segments_cross(p1, p2, p3, p4):
            return True, f"ตัดกับ TR_B edge ({u},{v})"

    return False, ""


def make_tie_line_feature(
    x1: float, y1: float, x2: float, y2: float,
    length_m_override: float = None,
    phase_designation: int = 7,
    conductor_size: int = 50,
) -> dict:
    length_m = (length_m_override if length_m_override is not None
                else math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2))
    return {
        "type": "Feature",
        "attributes": {
            "TAG": "LC_TIE",
            "SHAPE.LEN": length_m,
            "MEASURELENGTH": length_m,
            "CONDUCTORTYPE": "AW",
            "CONDUCTORSIZE": conductor_size,
            "PHASEDESIGNATION": phase_designation,
        },
        "geometry": {"paths": [[[x1, y1], [x2, y2]]]},
    }


def _tie_phase(net_a: "NetworkGraph", raw_a: dict, n_a: int) -> int:
    """
    Return PHASEDESIGNATION for the tie line based on the LC line feeding n_a.
    Uses the parent edge in TR_A's BFS tree so phantom phases are not created
    when n_a is on a single- or two-phase branch.
    """
    for parent in net_a.G.predecessors(n_a):
        for feat_idx, (fu, fv) in net_a.lc_feat_edges.items():
            if (fu == parent and fv == n_a) or (fu == n_a and fv == parent):
                pd = get_attr(raw_a["features"][feat_idx], "PHASEDESIGNATION", 7)
                return int(pd or 7)
    # Fallback: any LC line touching n_a
    for feat_idx, (fu, fv) in net_a.lc_feat_edges.items():
        if fu == n_a or fv == n_a:
            pd = get_attr(raw_a["features"][feat_idx], "PHASEDESIGNATION", 7)
            return int(pd or 7)
    return 7


def _nb_phase(net_b: "NetworkGraph", raw_b: dict, n_b: int) -> int:
    """
    Return the union of PHASEDESIGNATION bits for all LC lines touching n_b in TR_B.
    The tie line can only carry phases that n_b already has — using a wider phase
    would create floating nodes on the TR_B side and cause phantom voltages (~47V).
    """
    result = 0
    for feat_idx, (fu, fv) in net_b.lc_feat_edges.items():
        if fu == n_b or fv == n_b:
            pd = int(get_attr(raw_b["features"][feat_idx], "PHASEDESIGNATION", 7) or 7)
            result |= pd
    return result if result else 7


def _tie_conductor_size(net_a: "NetworkGraph", raw_a: dict, n_a: int,
                        net_b: "NetworkGraph", raw_b: dict, n_b: int) -> int:
    """
    Pick conductor size for the tie line: min of the size at n_a (TR_A parent edge)
    and the size at n_b (any LC line touching n_b in TR_B).
    Valid sizes: 50 or 95 (matches AW_IMP table).
    """
    def _size_at_node_parent(net, raw, node):
        for parent in net.G.predecessors(node):
            for feat_idx, (fu, fv) in net.lc_feat_edges.items():
                if (fu == parent and fv == node) or (fu == node and fv == parent):
                    sz = get_attr(raw["features"][feat_idx], "CONDUCTORSIZE", 50)
                    return int(sz or 50)
        for feat_idx, (fu, fv) in net.lc_feat_edges.items():
            if fu == node or fv == node:
                sz = get_attr(raw["features"][feat_idx], "CONDUCTORSIZE", 50)
                return int(sz or 50)
        return 50

    def _size_at_node_any(net, raw, node):
        for feat_idx, (fu, fv) in net.lc_feat_edges.items():
            if fu == node or fv == node:
                sz = get_attr(raw["features"][feat_idx], "CONDUCTORSIZE", 50)
                return int(sz or 50)
        return 50

    sz_a = _size_at_node_parent(net_a, raw_a, n_a)
    sz_b = _size_at_node_any(net_b, raw_b, n_b)
    raw_size = min(sz_a, sz_b)
    return 95 if raw_size >= 95 else 50


def _phase_restrict_feature(feat: dict, allowed_pd: int) -> Optional[dict]:
    """
    Return feat (or a copy) with PHASEDESIGNATION clipped to allowed_pd (bitwise AND).
    Returns None if the resulting phase is 0 (line entirely on excluded phases — skip it).
    Only call this on LC backbone line features, not on load/service features.
    """
    import copy
    orig_pd = int(get_attr(feat, "PHASEDESIGNATION", 7) or 7)
    new_pd = orig_pd & allowed_pd
    if new_pd == 0:
        return None
    if new_pd == orig_pd:
        return feat
    f = copy.deepcopy(feat)
    attrs = f.get("attributes") or {}
    attrs["PHASEDESIGNATION"] = new_pd
    return f


def _rephase_feature(feat: dict, new_pd: int) -> dict:
    """
    Return deep copy of feat with PHASEDESIGNATION and PHASE forced to new_pd.
    Used when transferring a subtree across a phase mismatch (TR_A phase ≠ TR_B phase).
    Never returns None.
    """
    import copy
    f = copy.deepcopy(feat)
    attrs = f.get("attributes") or {}
    orig_pd = int(attrs.get("PHASEDESIGNATION", 7) or 7)
    # Intersect if possible, otherwise force
    effective_pd = (orig_pd & new_pd) if (orig_pd & new_pd) else new_pd
    attrs["PHASEDESIGNATION"] = effective_pd
    # Update PHASE letter if the attribute exists
    if "PHASE" in attrs:
        if effective_pd & 4:
            attrs["PHASE"] = "A"
        elif effective_pd & 2:
            attrs["PHASE"] = "B"
        elif effective_pd & 1:
            attrs["PHASE"] = "C"
    return f


def query_transformer_loading(rated_kva: float) -> float:
    """Return loading % = actual secondary kVA / rated_kva * 100 from active circuit."""
    try:
        tx_names = odss.Transformers.AllNames()
        if not tx_names:
            return 0.0
        odss.Circuit.SetActiveElement(f"Transformer.{tx_names[0]}")
        pq_flat = odss.CktElement.Powers()           # [P0,Q0, P1,Q1, ...]
        pq = list(zip(pq_flat[0::2], pq_flat[1::2]))
        # Secondary terminal always has 4 conductors (phases 1,2,3 + neutral via .1.2.3.0).
        # Primary has 3 (delta) or 4 (wye) conductors depending on CONFIGURATION.
        # Use len(pq)-4 as offset so we always hit secondary phases 1,2,3.
        offset = len(pq) - 4
        P2 = sum(pq[offset + i][0] for i in range(min(3, len(pq) - offset)))
        Q2 = sum(pq[offset + i][1] for i in range(min(3, len(pq) - offset)))
        actual_kva = math.sqrt(P2 ** 2 + Q2 ** 2)
        return actual_kva / rated_kva * 100.0
    except Exception:
        return 0.0


def query_min_voltage_v() -> float:
    """Return minimum LV phase voltage (V) across backbone buses only (kVBase <= 1 kV),
    excluding meter buses (m_*) and neutrals so the value matches what is visible on the map."""
    try:
        min_v = float("inf")
        for busname in odss.Circuit.AllBusNames():
            if not busname:
                continue
            if busname.lower().startswith("m_"):   # skip meter/service buses
                continue
            odss.Circuit.SetActiveBus(busname)
            if odss.Bus.kVBase() > 1.0:   # skip HV buses (22 kV)
                continue
            nodes = list(odss.Bus.Nodes())
            vmag_ang = odss.Bus.VMagAngle()
            vmag = vmag_ang[0::2]
            for node, v in zip(nodes, vmag):
                fv = float(v)
                if node != 0 and fv > 5.0:
                    min_v = min(min_v, fv)
        return min_v if min_v < float("inf") else 0.0
    except Exception:
        return 0.0


def _collect_dss_bus_voltages() -> Dict[str, float]:
    """Return {bus_name: min_phase_V} for all LV buses in the active DSS circuit."""
    volt: Dict[str, float] = {}
    for busname in odss.Circuit.AllBusNames():
        if not busname:
            continue
        odss.Circuit.SetActiveBus(busname)
        if odss.Bus.kVBase() > 1.0:
            continue
        nodes = list(odss.Bus.Nodes())
        vmag = odss.Bus.VMagAngle()[0::2]
        valid = [float(v) for n, v in zip(nodes, vmag) if n != 0 and float(v) > 5.0]
        if valid:
            volt[busname] = min(valid)
    return volt


def _collect_dss_bus_phase_voltages() -> Dict[str, Dict[int, float]]:
    """Return {bus_name: {phase_node: V_float}} for all LV buses (node 1=A, 2=B, 3=C)."""
    result: Dict[str, Dict[int, float]] = {}
    for busname in odss.Circuit.AllBusNames():
        if not busname:
            continue
        odss.Circuit.SetActiveBus(busname)
        if odss.Bus.kVBase() > 1.0:
            continue
        nodes = list(odss.Bus.Nodes())
        vmag = odss.Bus.VMagAngle()[0::2]
        phases = {
            int(n): float(v)
            for n, v in zip(nodes, vmag)
            if int(n) != 0 and float(v) > 5.0
        }
        if phases:
            result[busname] = phases
    return result


def collect_optimal_node_voltages(
    net_a: "NetworkGraph",
    net_b: "NetworkGraph",
    raw_a: dict,
    raw_b: dict,
    opt: dict,
    snap_tol_a: float = 2.0,
    snap_tol_b: float = 2.0,
) -> Tuple[Dict[int, float], Dict[int, float],
           Dict[int, Dict[int, float]], Dict[int, Dict[int, float]]]:
    """
    Re-simulate the optimal transfer scenario.
    Returns (volt_a, volt_b, phase_a, phase_b):
      volt_a/b[node_id]        = min phase voltage (V)
      phase_a/b[node_id]       = {1:VA, 2:VB, 3:VC} (only phases present)
    Subtree nodes (drawn at TR_A coords, now fed by TR_B) → stored in volt_a/phase_a.
    """
    switch_edge = opt["switch_edge"]
    subtree     = opt["subtree_nodes"]
    n_a         = opt["tie_node_a"]
    n_b         = opt["tie_node_b"]

    remove_from_a: Set[int] = set(net_a.get_feature_indices_for_subtree(subtree))
    sw_feat = net_a.find_switch_edge_feature(switch_edge)
    if sw_feat is not None:
        remove_from_a.add(sw_feat)

    xa, ya = net_a.node_coords[n_a]
    xb, yb = net_b.node_coords[n_b]
    tie_dist_m = math.sqrt((xa - xb) ** 2 + (ya - yb) ** 2)
    na_pd = _tie_phase(net_a, raw_a, n_a)
    nb_pd = _nb_phase(net_b, raw_b, n_b)
    # Use TR_B's live phase as tie phase when no common phase exists
    tie_pd = na_pd & nb_pd if (na_pd & nb_pd) else nb_pd
    need_rephase = not bool(na_pd & nb_pd)
    tie_sz = _tie_conductor_size(net_a, raw_a, n_a, net_b, raw_b, n_b)
    tie_feature = make_tie_line_feature(
        xb, yb,
        xa + _COORD_OFFSET, ya + _COORD_OFFSET,
        length_m_override=tie_dist_m,
        phase_designation=tie_pd,
        conductor_size=tie_sz,
    )
    transferred = [tie_feature]
    for feat_idx in net_a.get_feature_indices_for_subtree(subtree):
        edge = net_a.lc_feat_edges.get(feat_idx)
        if edge is not None:
            u_e, v_e = edge
            if u_e in subtree and v_e in subtree:
                if need_rephase:
                    feat = _rephase_feature(raw_a["features"][feat_idx], tie_pd)
                else:
                    feat = _phase_restrict_feature(raw_a["features"][feat_idx], tie_pd)
                    if feat is None:
                        continue
                transferred.append(
                    _translate_feature(feat, _COORD_OFFSET, _COORD_OFFSET)
                )
        else:
            raw_feat = raw_a["features"][feat_idx]
            if need_rephase:
                raw_feat = _rephase_feature(raw_feat, tie_pd)
            transferred.append(
                _translate_feature(raw_feat, _COORD_OFFSET, _COORD_OFFSET)
            )

    volt_a:  Dict[int, float]           = {}
    volt_b:  Dict[int, float]           = {}
    phase_a: Dict[int, Dict[int, float]] = {}
    phase_b: Dict[int, Dict[int, float]] = {}

    tmp = tempfile.mkdtemp(prefix="topt_volt_")
    try:
        path_a = Path(tmp) / "mod_a.json"
        path_b = Path(tmp) / "mod_b.json"
        with open(path_a, "w", encoding="utf-8") as fh:
            json.dump(build_modified_json(raw_a, remove_from_a, []), fh, ensure_ascii=False)
        with open(path_b, "w", encoding="utf-8") as fh:
            json.dump(build_modified_json(raw_b, set(), transferred), fh, ensure_ascii=False)

        dss_a = str(Path(tmp) / "mod_a.dss")
        dss_b = str(Path(tmp) / "mod_b.dss")

        # Build KD-trees for coord → original node matching
        orig_a_ids  = list(net_a.node_coords.keys())
        orig_a_arr  = np.array([net_a.node_coords[i] for i in orig_a_ids])
        orig_a_tree = cKDTree(orig_a_arr)

        orig_b_ids  = list(net_b.node_coords.keys())
        orig_b_arr  = np.array([net_b.node_coords[i] for i in orig_b_ids])
        sub_a_ids   = list(subtree)
        sub_a_arr   = np.array([net_a.node_coords[i] for i in sub_a_ids])
        # TR_A subtree nodes are stored at offset coords in the modified TR_B JSON,
        # so build the search tree with the same offset applied.
        sub_a_arr_off = sub_a_arr + np.array([_COORD_OFFSET, _COORD_OFFSET])
        all_b_ids   = orig_b_ids + sub_a_ids
        all_b_arr   = np.vstack([orig_b_arr, sub_a_arr_off])
        orig_b_tree = cKDTree(all_b_arr)

        # ── TR_A modified ───────────────────────────────────────────────
        with contextlib.redirect_stdout(io.StringIO()):
            fa, _ = convert_json_to_dss_ordered(str(path_a), dss_a, snap_tol=snap_tol_a)
            solve_with_opendss(fa)
        all_dss_a = _collect_dss_bus_phase_voltages()
        print(
            f"[VOLTMAP] TR_A modified: converged={odss.Solution.Converged()} "
            f"buses={odss.Circuit.NumBuses()} bus_voltages_found={len(all_dss_a)}"
        )
        # Coordinate-based mapping: ใช้ SetBusXY ที่ DSS เขียนไว้ → ไม่ขึ้นกับ BFS ordering
        _n_phases = _n_zero_xy = _n_far = _n_matched = 0
        for bus in odss.Circuit.AllBusNames():
            phases = all_dss_a.get(bus, all_dss_a.get(bus.lower()))
            if not phases:
                continue
            _n_phases += 1
            odss.Circuit.SetActiveBus(bus)
            bx, by = odss.Bus.X(), odss.Bus.Y()
            if bx == 0.0 and by == 0.0:
                _n_zero_xy += 1
                continue
            dist, idx = orig_a_tree.query([bx, by])
            if dist < 100.0:
                _n_matched += 1
                nid = orig_a_ids[idx]
                v = min(phases.values())
                if nid not in volt_a or v < volt_a[nid]:
                    volt_a[nid] = v
                    if nid not in phase_a or v == volt_a[nid]:
                        phase_a[nid] = phases
            else:
                _n_far += 1
        print(
            f"[VOLTMAP] TR_A match: phases_ok={_n_phases} zero_xy={_n_zero_xy} "
            f"too_far={_n_far} matched={_n_matched} volt_a_size={len(volt_a)}"
        )

        # ── TR_B modified ───────────────────────────────────────────────
        with contextlib.redirect_stdout(io.StringIO()):
            fb, _ = convert_json_to_dss_ordered(str(path_b), dss_b, snap_tol=snap_tol_b)
            solve_with_opendss(fb)
        n_orig_b = len(orig_b_ids)
        all_dss_b = _collect_dss_bus_phase_voltages()
        print(
            f"[VOLTMAP] TR_B modified: converged={odss.Solution.Converged()} "
            f"buses={odss.Circuit.NumBuses()} bus_voltages_found={len(all_dss_b)}"
        )
        # Coordinate-based mapping สำหรับ TR_B (รวม transferred subtree ที่ offset ไปแล้ว)
        _n_phases_b = _n_zero_xy_b = _n_far_b = _n_matched_b = 0
        for bus in odss.Circuit.AllBusNames():
            phases = all_dss_b.get(bus, all_dss_b.get(bus.lower()))
            if not phases:
                continue
            _n_phases_b += 1
            odss.Circuit.SetActiveBus(bus)
            bx, by = odss.Bus.X(), odss.Bus.Y()
            if bx == 0.0 and by == 0.0:
                _n_zero_xy_b += 1
                continue
            dist, idx = orig_b_tree.query([bx, by])
            if dist < 100.0:
                _n_matched_b += 1
                v = min(phases.values())
                if idx < n_orig_b:
                    nid = orig_b_ids[idx]
                    if nid not in volt_b or v < volt_b[nid]:
                        volt_b[nid] = v
                        if nid not in phase_b or v == volt_b[nid]:
                            phase_b[nid] = phases
                else:
                    nid = sub_a_ids[idx - n_orig_b]
                    if nid not in volt_a or v < volt_a[nid]:
                        volt_a[nid] = v
                        if nid not in phase_a or v == volt_a[nid]:
                            phase_a[nid] = phases
            else:
                _n_far_b += 1
        print(
            f"[VOLTMAP] TR_B match: phases_ok={_n_phases_b} zero_xy={_n_zero_xy_b} "
            f"too_far={_n_far_b} matched={_n_matched_b} volt_b_size={len(volt_b)}"
        )

        # ── Meter voltages by bus name ─────────────────────────────────
        # The coordinate matching above needs SetBusXY-written bus
        # coordinates, which convert_json_to_dss_ordered never writes (see
        # zero_xy counts) — so it can never match anything. Meters don't
        # need it: convert_json_to_dss_ordered names each meter's bus
        # deterministically as "M_<PEANO>" (build_meter_bus_names replicates
        # that naming exactly), so look voltages up by that name instead.
        # Falls back across both circuits since a transferred subtree's
        # meters end up simulated in TR_B's circuit even though they're
        # drawn/reported at TR_A's node ids.
        meter_bus_a = build_meter_bus_names(raw_a["features"])
        meter_bus_b = build_meter_bus_names(raw_b["features"])

        _n_meter_a = _n_meter_a_hit = 0
        for feat_idx, nid in net_a.load_feat_nodes.items():
            bus_name = meter_bus_a.get(feat_idx)
            if not bus_name:
                continue
            _n_meter_a += 1
            key = bus_name.lower()
            phases = (all_dss_a.get(key) or all_dss_a.get(bus_name)
                      or all_dss_b.get(key) or all_dss_b.get(bus_name))
            if not phases:
                continue
            _n_meter_a_hit += 1
            v = min(phases.values())
            if nid not in volt_a or v < volt_a[nid]:
                volt_a[nid] = v
                if nid not in phase_a or v == volt_a[nid]:
                    phase_a[nid] = phases

        _n_meter_b = _n_meter_b_hit = 0
        for feat_idx, nid in net_b.load_feat_nodes.items():
            bus_name = meter_bus_b.get(feat_idx)
            if not bus_name:
                continue
            _n_meter_b += 1
            key = bus_name.lower()
            phases = all_dss_b.get(key) or all_dss_b.get(bus_name)
            if not phases:
                continue
            _n_meter_b_hit += 1
            v = min(phases.values())
            if nid not in volt_b or v < volt_b[nid]:
                volt_b[nid] = v
                if nid not in phase_b or v == volt_b[nid]:
                    phase_b[nid] = phases

        print(
            f"[VOLTMAP] Meter match by name: TR_A {_n_meter_a_hit}/{_n_meter_a}  "
            f"TR_B {_n_meter_b_hit}/{_n_meter_b}"
        )

    except Exception as exc:
        print(f"[WARN] voltage map collection failed: {exc}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    # Propagate voltages from connected backbone nodes to island nodes (no DSS voltage).
    # Island nodes exist when LC lines in the JSON are disconnected from the transformer.
    # Customers snapping to island nodes otherwise show gray (no voltage) in the map.
    for net, volt, phase in [(net_a, volt_a, phase_a), (net_b, volt_b, phase_b)]:
        missing_ids = [nid for nid in net.node_coords if nid not in volt]
        if not missing_ids or not volt:
            continue
        conn_ids = [nid for nid in volt if nid in net.node_coords]
        if not conn_ids:
            continue
        conn_arr = np.array([net.node_coords[i] for i in conn_ids])
        conn_tree = cKDTree(conn_arr)
        for nid in missing_ids:
            coords = net.node_coords.get(nid)
            if coords is None:
                continue
            _, idx = conn_tree.query(list(coords))
            volt[nid]  = volt[conn_ids[idx]]
            phase[nid] = phase.get(conn_ids[idx], {})

    return volt_a, volt_b, phase_a, phase_b


# =====================================================================
# TransferOptimizer
# =====================================================================

class TransferOptimizer:
    def __init__(
        self,
        fac_a: str,
        fac_b: str,
        max_tie_dist: float = 100.0,
        load_limit_pct: float = 80.0,
        min_voltage_v: float = 210.0,
        json_dir: str = OUTPUT_DIR,
        force_refresh: bool = False,
        region: Optional[str] = None,
    ) -> None:
        self.fac_a = fac_a
        self.fac_b = fac_b
        self.max_tie_dist = max_tie_dist
        self.load_limit_pct = load_limit_pct
        self.min_voltage_v = min_voltage_v
        self.json_dir = json_dir
        self.force_refresh = force_refresh
        # PEA GIS region (e.g. "NE1", "C2", "Z") both facilities live in — needed
        # to point InputJsonApi at the right regional GIS server before querying;
        # without it, InputJsonApi silently defaults to NE1 and any facility in
        # another region fails with "ไม่พบ TR สำหรับ FACILITYID=...".
        self.region = region

        self.net_a: Optional[NetworkGraph] = None
        self.net_b: Optional[NetworkGraph] = None
        self._json_a: Optional[Path] = None
        self._json_b: Optional[Path] = None
        self.a_baseline_pct: float = 0.0
        self.b_baseline_pct: float = 0.0
        self.a_baseline_minv: float = 0.0
        self.b_baseline_minv: float = 0.0

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _ensure_json(self, fac: str) -> Path:
        # Legacy on-disk cache locations (self.json_dir, e.g. D:\testpy\jsonfile) —
        # kept as a fast path in case files were ever pre-placed there.
        candidates = [
            Path(self.json_dir) / f"NetworkLV{fac}_with_MV.json",
            Path(self.json_dir) / f"NetworkLV_{fac}_with_MV.json",
            Path(self.json_dir) / f"NetworkLV{fac}.json",
            Path(self.json_dir) / f"NetworkLV_{fac}.json",
        ]
        if not self.force_refresh:
            for p in candidates:
                if p.exists():
                    print(f"  [JSON] พบ cache: {p.name}")
                    return p.resolve()

        import InputJsonApi as inf
        # Point InputJsonApi at the right regional GIS server first — it defaults
        # to NE1 otherwise, so any facility outside NE1 would fail to be found.
        inf.set_gis_region(self.region)
        print(f"  [JSON] ดึงข้อมูลจาก API FACILITYID: {fac} (region={self.region})...")
        # InputJsonApi writes to pea_no_projects/input/<project_id>/... (relative
        # to cwd) and returns that exact path — use it directly instead of
        # guessing a location, since it doesn't write into self.json_dir.
        # Resolve to absolute immediately: OpenDSS's Compile command changes
        # the process cwd as a side effect (see solve_with_opendss), so a
        # relative path cached here would silently break on the next facility.
        result = inf.run_once_with_facilityid(fac, project_id=f"shareload_{fac}")
        out_path = result.get("out_path")
        if out_path and Path(out_path).exists():
            return Path(out_path).resolve()

        for p in candidates:
            if p.exists():
                return p.resolve()
        raise FileNotFoundError(f"ไม่พบ JSON สำหรับ {fac} แม้หลังดึงข้อมูลจาก API")

    def _run_baseline_sim(self, json_path: Path, rated_kva: float, label: str,
                          snap_tol: float = 2.0) -> Tuple[float, float]:
        """Run OpenDSS baseline for one network. Returns (loading_pct, min_voltage_v)."""
        tmp = tempfile.mkdtemp(prefix="topt_base_")
        try:
            dss_path = str(Path(tmp) / f"base_{label}.dss")
            dss_file, _ = convert_json_to_dss_ordered(str(json_path), dss_path, snap_tol=snap_tol)
            solve_with_opendss(dss_file)
            loading = query_transformer_loading(rated_kva)
            min_v = query_min_voltage_v()
            return loading, min_v
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def _simulate_scenario(
        self,
        scenario: dict,
        raw_a: dict,
        raw_b: dict,
        snap_tol_a: float = 2.0,
        snap_tol_b: float = 2.0,
    ) -> dict:
        """Build modified JSONs, run both sims, record constraint check."""
        switch_edge = scenario["switch_edge"]
        subtree = scenario["subtree_nodes"]
        n_a = scenario["tie_node_a"]
        n_b = scenario["tie_node_b"]

        # Features to remove from TR_A:
        # 1. Subtree-internal lines + loads + service lines
        # 2. The switch-edge line feature itself
        remove_from_a: Set[int] = set(self.net_a.get_feature_indices_for_subtree(subtree))
        sw_feat = self.net_a.find_switch_edge_feature(switch_edge)
        if sw_feat is not None:
            remove_from_a.add(sw_feat)

        # Features to add to TR_B:
        # 1. Tie line (n_b coord → n_a coord, with TR_A end offset to prevent cross-network snapping)
        # 2. Subtree-internal lines (both endpoints in subtree, coordinates offset)
        # 3. Loads and service lines from subtree (coordinates offset)
        #
        # The _COORD_OFFSET shifts all TR_A subtree coordinates into a "virtual" space so
        # cluster_points inside convert_json_to_dss_ordered never merges TR_A and TR_B nodes
        # that happen to be physically close (e.g. tie_dist < snap_tol=2m).
        xa, ya = self.net_a.node_coords[n_a]
        xb, yb = self.net_b.node_coords[n_b]
        na_pd = _tie_phase(self.net_a, raw_a, n_a)
        nb_pd = _nb_phase(self.net_b, raw_b, n_b)
        # Use TR_B's live phase as tie phase when no common phase exists
        tie_pd = na_pd & nb_pd if (na_pd & nb_pd) else nb_pd
        need_rephase = not bool(na_pd & nb_pd)
        tie_sz = _tie_conductor_size(self.net_a, raw_a, n_a, self.net_b, raw_b, n_b)
        tie_feature = make_tie_line_feature(
            xb, yb,
            xa + _COORD_OFFSET, ya + _COORD_OFFSET,
            length_m_override=scenario["tie_distance_m"],
            phase_designation=tie_pd,
            conductor_size=tie_sz,
        )

        transferred_features = [tie_feature]
        for feat_idx in self.net_a.get_feature_indices_for_subtree(subtree):
            edge = self.net_a.lc_feat_edges.get(feat_idx)
            if edge is not None:
                u_e, v_e = edge
                if u_e in subtree and v_e in subtree:
                    if need_rephase:
                        feat = _rephase_feature(raw_a["features"][feat_idx], tie_pd)
                    else:
                        feat = _phase_restrict_feature(raw_a["features"][feat_idx], tie_pd)
                        if feat is None:
                            continue
                    transferred_features.append(
                        _translate_feature(feat, _COORD_OFFSET, _COORD_OFFSET)
                    )
            else:
                raw_feat = raw_a["features"][feat_idx]
                if need_rephase:
                    raw_feat = _rephase_feature(raw_feat, tie_pd)
                transferred_features.append(
                    _translate_feature(raw_feat, _COORD_OFFSET, _COORD_OFFSET)
                )

        tmp = tempfile.mkdtemp(prefix="topt_scen_")
        result = dict(scenario)
        result["tie_conductor_size"] = tie_sz
        result.setdefault("error", None)
        try:
            path_a = Path(tmp) / "mod_a.json"
            path_b = Path(tmp) / "mod_b.json"
            with open(path_a, "w", encoding="utf-8") as fh:
                json.dump(build_modified_json(raw_a, remove_from_a, []), fh, ensure_ascii=False)
            with open(path_b, "w", encoding="utf-8") as fh:
                json.dump(build_modified_json(raw_b, set(), transferred_features), fh, ensure_ascii=False)

            dss_a = str(Path(tmp) / "mod_a.dss")
            dss_b = str(Path(tmp) / "mod_b.dss")

            b_baseline_pct = scenario.get("b_loading_before", 0.0)

            # Simulate modified TR_A
            try:
                fa, _ = convert_json_to_dss_ordered(str(path_a), dss_a, snap_tol=snap_tol_a)
                solve_with_opendss(fa)
                if not odss.Solution.Converged():
                    result["error"] = "TR_A sim: not converged"
                    result["feasible"] = False
                    return result
                a_loading = query_transformer_loading(self.net_a.rated_kva)
                a_min_v = query_min_voltage_v()
                # Sanity: degenerate convergence gives huge loading
                if a_loading > 500.0:
                    result["error"] = f"TR_A sim: loading {a_loading:.0f}% (degenerate)"
                    result["feasible"] = False
                    return result
            except Exception as exc:
                result["error"] = f"TR_A sim: {exc}"
                result["feasible"] = False
                return result

            # Simulate modified TR_B (extended)
            try:
                fb, _ = convert_json_to_dss_ordered(str(path_b), dss_b, snap_tol=snap_tol_b)
                solve_with_opendss(fb)
                if not odss.Solution.Converged():
                    result["error"] = "TR_B sim: not converged"
                    result["feasible"] = False
                    return result
                b_loading = query_transformer_loading(self.net_b.rated_kva)
                b_min_v = query_min_voltage_v()
                # Sanity: if TR_B had significant load before but now shows near-zero,
                # the transferred subtree is floating (degenerate convergence).
                if b_loading < 2.0 and b_baseline_pct > 10.0:
                    result["error"] = (
                        f"TR_B sim: loading {b_loading:.1f}% despite baseline "
                        f"{b_baseline_pct:.1f}% (degenerate)"
                    )
                    result["feasible"] = False
                    return result
            except Exception as exc:
                result["error"] = f"TR_B sim: {exc}"
                result["feasible"] = False
                return result

            result["a_loading_after"] = a_loading
            result["b_loading_after"] = b_loading
            result["a_min_v"] = a_min_v
            result["b_min_v"] = b_min_v
            result["min_v"] = min(a_min_v, b_min_v)
            # DBG: print which bus gives min voltage in TR_B simulation
            _dbg_min = float("inf"); _dbg_bus = "?"
            for _bn in odss.Circuit.AllBusNames():
                if not _bn: continue
                odss.Circuit.SetActiveBus(_bn)
                if odss.Bus.kVBase() > 1.0: continue
                for _nd, _v in zip(odss.Bus.Nodes(), odss.Bus.VMagAngle()[0::2]):
                    if int(_nd) != 0 and float(_v) > 5.0 and float(_v) < _dbg_min:
                        _dbg_min = float(_v); _dbg_bus = f"{_bn}.{_nd}  kVBase={odss.Bus.kVBase():.4f}"
            print(f"    [DBG] a={a_min_v:.1f}V b={b_min_v:.1f}V  min_bus={_dbg_bus}")
            result["feasible"] = (
                a_loading < self.load_limit_pct
                and b_loading < self.load_limit_pct
                and result["min_v"] >= self.min_voltage_v
            )
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

        return result

    # ------------------------------------------------------------------
    # Main entry
    # ------------------------------------------------------------------

    def run(self) -> List[dict]:
        """
        Execute the full optimization pipeline.
        Returns list of scenario dicts (all candidates; feasible ones have feasible=True).
        Optimal solution is marked with optimal=True.
        """
        print(f"\n{'='*60}")
        print(f"Transformer Neighborhood Optimizer")
        print(f"  TR_A (overloaded) : {self.fac_a}")
        print(f"  TR_B (available)  : {self.fac_b}")
        print(f"  max_tie_dist      : {self.max_tie_dist} m")
        print(f"  load_limit        : {self.load_limit_pct}%")
        print(f"  min_voltage       : {self.min_voltage_v} V")
        print(f"{'='*60}")

        # Step 1: Ensure JSON files exist
        print("\n[Step 1] Import LV Network (JSON files)...")
        self._json_a = self._ensure_json(self.fac_a)
        self._json_b = self._ensure_json(self.fac_b)

        # Step 2: Parse both networks
        print("\n[Step 2] Parse networks...")
        self.net_a = NetworkGraph(str(self._json_a))
        print(f"  TR_A: {len(self.net_a.G.nodes())} nodes, {len(self.net_a.G.edges())} edges, "
              f"rated={self.net_a.rated_kva:.0f} kVA, total_kW={sum(self.net_a.node_kw.values()):.1f}")
        self.net_b = NetworkGraph(str(self._json_b))
        print(f"  TR_B: {len(self.net_b.G.nodes())} nodes, {len(self.net_b.G.edges())} edges, "
              f"rated={self.net_b.rated_kva:.0f} kVA, total_kW={sum(self.net_b.node_kw.values()):.1f}")

        # Load raw JSONs for modification
        with open(self._json_a, encoding="utf-8") as fh:
            raw_a = json.load(fh)
        with open(self._json_b, encoding="utf-8") as fh:
            raw_b = json.load(fh)

        # Step 3: Baseline simulations (pass detected snap_tol so DSS sees same topology)
        snap_tol_a = self.net_a.snap_tol
        snap_tol_b = self.net_b.snap_tol
        print("\n[Step 3] Baseline OpenDSS simulation...")
        a_baseline, a_base_minv = self._run_baseline_sim(
            self._json_a, self.net_a.rated_kva, "A", snap_tol=snap_tol_a)
        print(f"  TR_A: loading={a_baseline:.1f}%  V_min={a_base_minv:.1f}V")
        b_baseline, b_base_minv = self._run_baseline_sim(
            self._json_b, self.net_b.rated_kva, "B", snap_tol=snap_tol_b)
        print(f"  TR_B: loading={b_baseline:.1f}%  V_min={b_base_minv:.1f}V")
        self.a_baseline_pct  = a_baseline
        self.b_baseline_pct  = b_baseline
        self.a_baseline_minv = a_base_minv
        self.b_baseline_minv = b_base_minv

        if a_baseline < self.load_limit_pct:
            print(f"  [!] TR_A loading={a_baseline:.1f}% < {self.load_limit_pct}% (ต่ำกว่าเกณฑ์)")
        if b_baseline >= self.load_limit_pct:
            print(f"  [!] TR_B loading={b_baseline:.1f}% >= {self.load_limit_pct}% (ไม่มีพื้นที่รับโหลด)")

        # Step 4: Build cKDTree of TR_B nodes
        b_node_ids = list(self.net_b.node_coords.keys())
        b_coords_arr = np.array([self.net_b.node_coords[n] for n in b_node_ids])
        b_tree = cKDTree(b_coords_arr)

        # Step 5: Find candidate scenarios
        print("\n[Step 4] หา candidate scenarios (switch + tie)...")
        candidates = []
        for (u, v) in list(self.net_a.G.edges()):
            subtree = self.net_a.get_subtree((u, v))
            if not subtree:
                continue
            subtree_kw = self.net_a.get_total_kw(subtree)
            if subtree_kw <= 0.0:
                continue

            # Nearest TR_B node to any node in this subtree
            best_dist = float("inf")
            best_na = -1
            best_nb_idx = -1
            for n_a in subtree:
                xy = self.net_a.node_coords[n_a]
                dist, idx = b_tree.query([xy[0], xy[1]], k=1)
                if dist < best_dist:
                    best_dist = dist
                    best_na = n_a
                    best_nb_idx = int(idx)

            if best_dist > self.max_tie_dist:
                continue

            best_nb = b_node_ids[best_nb_idx]

            # Physical feasibility: tie must not cross existing LV segments
            crossed, cross_desc = _tie_crosses_lv(best_na, best_nb, self.net_a, self.net_b)
            if crossed:
                print(f"  [PHYS-NG] switch=({u},{v}) tie {best_na}→{best_nb}  {cross_desc}")
                continue

            # Estimate post-transfer loading (kW-based estimate)
            est_a_kva = max(0.0, (a_baseline / 100.0) * self.net_a.rated_kva - subtree_kw)
            est_b_kva = (b_baseline / 100.0) * self.net_b.rated_kva + subtree_kw
            est_a_pct = est_a_kva / self.net_a.rated_kva * 100.0
            est_b_pct = est_b_kva / self.net_b.rated_kva * 100.0
            if est_a_pct >= self.load_limit_pct or est_b_pct >= self.load_limit_pct:
                continue

            candidates.append({
                "switch_edge": (u, v),
                "tie_node_a": best_na,
                "tie_node_b": best_nb,
                "tie_distance_m": best_dist,
                "subtree_nodes": subtree,
                "subtree_kw": subtree_kw,
                "a_loading_before": a_baseline,
                "b_loading_before": b_baseline,
                "est_a_pct": est_a_pct,
                "est_b_pct": est_b_pct,
                "feasible": False,
                "optimal": False,
                "error": None,
            })

        print(f"  พบ {len(candidates)} candidate scenarios")
        if not candidates:
            print("  [!] ไม่พบ scenario ที่ประมาณการว่า feasible")
            print("      ลองเพิ่ม --max-tie-dist หรือตรวจสอบว่าทั้ง 2 เครือข่ายอยู่ใกล้กัน")
            return []

        # Step 6: Simulate each candidate
        print("\n[Step 5] รัน OpenDSS simulation แต่ละ scenario...")
        results: List[dict] = []
        for i, scen in enumerate(candidates):
            su, sv = scen["switch_edge"]
            print(f"  [{i+1:>3}/{len(candidates)}] switch=({su},{sv})  "
                  f"tie_dist={scen['tie_distance_m']:.1f}m  kW={scen['subtree_kw']:.1f}", end="  ")
            r = self._simulate_scenario(scen, raw_a, raw_b,
                                        snap_tol_a=snap_tol_a, snap_tol_b=snap_tol_b)
            results.append(r)
            if r.get("feasible"):
                print(f"[OK]  A={r['a_loading_after']:.1f}%  B={r['b_loading_after']:.1f}%  "
                      f"V_min={r['min_v']:.0f}V")
            else:
                if r.get("error"):
                    err_msg = r["error"]
                elif "a_loading_after" in r:
                    err_msg = (f"A={r['a_loading_after']:.1f}%  "
                               f"B={r['b_loading_after']:.1f}%  "
                               f"V={r['min_v']:.0f}V")
                else:
                    err_msg = "sim failed"
                print(f"[NG]  {err_msg}")

        # Step 7: Mark optimal
        feasible = [r for r in results if r.get("feasible")]
        if feasible:
            # Primary: maximize V_min (rounded to 0.1V); Secondary: minimize A_after%; Tertiary: minimize tie distance
            feasible.sort(key=lambda r: (
                -round(r.get("min_v", 0.0), 1),
                r["a_loading_after"],
                r.get("tie_distance_m", 0.0),
            ))
            feasible[0]["optimal"] = True

        n_feasible = len(feasible)
        print(f"\n  Feasible scenarios: {n_feasible} / {len(results)}")
        if n_feasible > 0:
            opt = feasible[0]
            print(f"  Optimal: switch=({opt['switch_edge'][0]},{opt['switch_edge'][1]})  "
                  f"kW_transferred={opt['subtree_kw']:.1f}  "
                  f"A={opt['a_loading_after']:.1f}%  B={opt['b_loading_after']:.1f}%  "
                  f"V_min={opt['min_v']:.0f}V")

        return results


# =====================================================================
# Output: console table
# =====================================================================

def print_results_table(results: List[dict], net_a: NetworkGraph, net_b: NetworkGraph) -> None:
    if not results:
        return
    base_a = results[0].get("a_loading_before", 0.0)
    base_b = results[0].get("b_loading_before", 0.0)
    col = 110
    print(f"\n{'='*col}")
    print(f"  LOAD TRANSFER RESULTS  |  {net_a.facilityid} → {net_b.facilityid}")
    print(f"  Baseline: TR_A={base_a:.1f}%  TR_B={base_b:.1f}%  "
          f"(Rated: A={net_a.rated_kva:.0f}kVA  B={net_b.rated_kva:.0f}kVA)")
    print(f"{'='*col}")
    hdr = (f"{'#':>4} {'SW_U':>8} {'SW_V':>8} {'TieA':>8} {'TieB':>8} "
           f"{'Dist(m)':>8} {'kW_xfr':>8} {'A_bef%':>7} {'A_aft%':>7} "
           f"{'B_bef%':>7} {'B_aft%':>7} {'V_min(V)':>9} {'CndSz':>6} {'OK':>5}")
    print(hdr)
    print("-" * col)
    rank = 0
    for r in results:
        if not r.get("feasible"):
            continue
        rank += 1
        mark = "[OPT]" if r.get("optimal") else "[OK]"
        su, sv = r["switch_edge"]
        na, nb = r["tie_node_a"], r["tie_node_b"]
        csz = r.get("tie_conductor_size", 50)
        row = (f"{rank:>4} {su:>8} {sv:>8} {na:>8} {nb:>8} "
               f"{r['tie_distance_m']:>8.1f} {r['subtree_kw']:>8.1f} "
               f"{r['a_loading_before']:>7.1f} {r['a_loading_after']:>7.1f} "
               f"{r['b_loading_before']:>7.1f} {r['b_loading_after']:>7.1f} "
               f"{r['min_v']:>9.1f} {csz:>4}mm2 {mark:>5}")
        print(row)
    if rank == 0:
        print("  (ไม่พบ scenario ที่ผ่านเงื่อนไข -- ลองปรับ --max-tie-dist หรือ --load-limit)")
    print("=" * col)


# =====================================================================
# Output: Excel report
# =====================================================================

def save_excel_report(
    results: List[dict],
    net_a: NetworkGraph,
    net_b: NetworkGraph,
    out_path: str,
) -> None:
    wb = openpyxl.Workbook()

    # ---- Sheet: Results ----
    ws = wb.active
    ws.title = "Results"
    hdr_fill = PatternFill("solid", fgColor="1F497D")
    hdr_font = Font(color="FFFFFF", bold=True)
    opt_fill = PatternFill("solid", fgColor="E2EFDA")
    bad_fill = PatternFill("solid", fgColor="FCE4D6")

    headers = [
        "Rank", "Status",
        "Switch_Parent_Node", "Switch_Child_Node",
        "TieNode_A", "TieNode_B",
        "TieDistance_m", "TransferredLoad_kW",
        "TieConductorSize_mm2",
        "TRA_Loading_Before%", "TRB_Loading_Before%",
        "TRA_Loading_After%", "TRB_Loading_After%",
        "TRA_MinVolt_V", "TRB_MinVolt_V", "Overall_MinVolt_V",
        "Feasible", "Error",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    rank = 0
    for r in results:
        feasible = r.get("feasible", False)
        if feasible:
            rank += 1
        su, sv = r["switch_edge"]
        na, nb = r.get("tie_node_a", ""), r.get("tie_node_b", "")
        status = "** OPTIMAL **" if r.get("optimal") else ("OK" if feasible else "NG")
        row = [
            rank if feasible else "",
            status,
            su, sv, na, nb,
            round(r.get("tie_distance_m", 0), 1),
            round(r.get("subtree_kw", 0), 2),
            r.get("tie_conductor_size", 50),
            round(r.get("a_loading_before", 0), 1),
            round(r.get("b_loading_before", 0), 1),
            round(r.get("a_loading_after", 0), 1) if "a_loading_after" in r else "",
            round(r.get("b_loading_after", 0), 1) if "b_loading_after" in r else "",
            round(r.get("a_min_v", 0), 1) if "a_min_v" in r else "",
            round(r.get("b_min_v", 0), 1) if "b_min_v" in r else "",
            round(r.get("min_v", 0), 1) if "min_v" in r else "",
            "YES" if feasible else "NO",
            r.get("error") or "",
        ]
        ws.append(row)
        ridx = ws.max_row
        fill = opt_fill if r.get("optimal") else (None if feasible else bad_fill)
        if fill:
            for cell in ws[ridx]:
                cell.fill = fill

    for col in ws.columns:
        max_w = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(10, max_w + 2), 32)

    # ---- Sheet: Summary ----
    ws2 = wb.create_sheet("Summary")
    feasible_list = [r for r in results if r.get("feasible")]
    opt = next((r for r in results if r.get("optimal")), None)

    rows_summary = [
        ["Parameter", "Value"],
        ["TR_A (overloaded)", net_a.facilityid],
        ["TR_B (available)", net_b.facilityid],
        ["TR_A Rated kVA", net_a.rated_kva],
        ["TR_B Rated kVA", net_b.rated_kva],
        ["TR_A Baseline Loading%", round(results[0]["a_loading_before"], 1) if results else ""],
        ["TR_B Baseline Loading%", round(results[0]["b_loading_before"], 1) if results else ""],
        ["Total Scenarios Evaluated", len(results)],
        ["Feasible Scenarios", len(feasible_list)],
    ]
    if opt:
        rows_summary += [
            ["", ""],
            ["--- Optimal Solution ---", ""],
            ["Switch Parent Node", opt["switch_edge"][0]],
            ["Switch Child Node", opt["switch_edge"][1]],
            ["Tie Node (TR_A side)", opt["tie_node_a"]],
            ["Tie Node (TR_B side)", opt["tie_node_b"]],
            ["Tie Line Distance (m)", round(opt["tie_distance_m"], 1)],
            ["Load Transferred (kW)", round(opt["subtree_kw"], 2)],
            ["TR_A Loading After (%)", round(opt["a_loading_after"], 1)],
            ["TR_B Loading After (%)", round(opt["b_loading_after"], 1)],
            ["Minimum Voltage (V)", round(opt["min_v"], 1)],
        ]
    for row in rows_summary:
        ws2.append(row)

    wb.save(out_path)
    print(f"[Excel] บันทึก: {out_path}")


# =====================================================================
# Output: topology map
# =====================================================================

def draw_topology_map(
    net_a: NetworkGraph,
    net_b: NetworkGraph,
    results: List[dict],
    out_path: str,
    volt_a: Dict[int, float] = None,
    volt_b: Dict[int, float] = None,
) -> None:
    volt_a = volt_a or {}
    volt_b = volt_b or {}
    has_volt = bool(volt_a or volt_b)

    norm     = mcolors.Normalize(vmin=150, vmax=240)
    cmap     = mcolors.LinearSegmentedColormap.from_list(
        "volt", ["#d73027","#fee08b","#91cf60","#1a9850"], N=256)
    DEF_A    = "#4472C4"
    DEF_B    = "#ED7D31"

    fig, ax = plt.subplots(figsize=(16, 11))

    # --- Edges ---
    for u, v in net_a.G.edges():
        xu, yu = net_a.node_coords[u]; xv, yv = net_a.node_coords[v]
        ax.plot([xu, xv], [yu, yv], color=DEF_A, linewidth=1.5, alpha=0.45, zorder=2)
    for u, v in net_b.G.edges():
        xu, yu = net_b.node_coords[u]; xv, yv = net_b.node_coords[v]
        ax.plot([xu, xv], [yu, yv], color=DEF_B, linewidth=1.5, alpha=0.45, zorder=2)

    # --- Nodes (color = voltage if available, else network color) ---
    for nid, (x, y) in net_a.node_coords.items():
        s = 30 + net_a.node_kw.get(nid, 0) * 1.5
        c = cmap(norm(volt_a[nid])) if nid in volt_a else DEF_A
        ax.scatter(x, y, color=c, s=s, zorder=3, alpha=0.88)
    for nid, (x, y) in net_b.node_coords.items():
        s = 30 + net_b.node_kw.get(nid, 0) * 1.5
        c = cmap(norm(volt_b[nid])) if nid in volt_b else DEF_B
        ax.scatter(x, y, color=c, s=s, zorder=3, alpha=0.88)

    # --- Red ring for nodes < 200V ---
    for nid, v in volt_a.items():
        if v < 200:
            x, y = net_a.node_coords[nid]
            ax.scatter(x, y, s=90, facecolors="none", edgecolors="red", linewidths=1.5, zorder=4)
    for nid, v in volt_b.items():
        if v < 200:
            x, y = net_b.node_coords[nid]
            ax.scatter(x, y, s=90, facecolors="none", edgecolors="red", linewidths=1.5, zorder=4)

    # --- Transformer nodes ---
    tx_ax, tx_ay = net_a.node_coords[net_a.transformer_node]
    tx_bx, tx_by = net_b.node_coords[net_b.transformer_node]
    ax.scatter(tx_ax, tx_ay, color=DEF_A, marker="*", s=500, zorder=5, edgecolors="k", linewidths=0.8)
    ax.scatter(tx_bx, tx_by, color=DEF_B, marker="*", s=500, zorder=5, edgecolors="k", linewidths=0.8)
    ax.annotate(f"TR_A\n{net_a.facilityid}", (tx_ax, tx_ay),
                xytext=(6, 6), textcoords="offset points", fontsize=8, color=DEF_A, fontweight="bold")
    ax.annotate(f"TR_B\n{net_b.facilityid}", (tx_bx, tx_by),
                xytext=(6, 6), textcoords="offset points", fontsize=8, color=DEF_B, fontweight="bold")

    # --- Non-optimal feasible switch edges (thin dotted, labeled with table row #) ---
    for idx, r in enumerate(results, start=1):
        if not r.get("feasible") or r.get("optimal"):
            continue
        su2, sv2 = r["switch_edge"]
        xu2, yu2 = net_a.node_coords[su2]; xv2, yv2 = net_a.node_coords[sv2]
        ax.plot([xu2, xv2], [yu2, yv2], color="#FF9999", linewidth=1.5,
                linestyle=":", zorder=5, alpha=0.8)
        mx2, my2 = (xu2 + xv2) / 2, (yu2 + yv2) / 2
        ax.text(mx2, my2, f"#{idx}", fontsize=6, ha="center", va="center",
                color="crimson", zorder=6,
                bbox=dict(fc="white", ec="#FF9999", pad=1, alpha=0.7, boxstyle="round,pad=0.2"))

    # --- Optimal scenario ---
    opt = next((r for r in results if r.get("optimal")), None)
    opt_idx = next((i for i, r in enumerate(results, start=1) if r.get("optimal")), None)
    if opt:
        su, sv = opt["switch_edge"]
        xu, yu = net_a.node_coords[su]; xv, yv = net_a.node_coords[sv]

        # Switch edge
        ax.plot([xu, xv], [yu, yv], color="red", linewidth=3, linestyle="--", zorder=6)
        ax.scatter([xu, xv], [yu, yv], color="red", s=160, zorder=7, marker="x", linewidths=2.5)

        # Switch node labels with voltage
        v_su = volt_a.get(su); v_sv = volt_a.get(sv)
        lbl_su = f"sw_u  nid={su}" + (f"\n{v_su:.0f} V" if v_su else "")
        lbl_sv = f"sw_v  nid={sv}" + (f"\n{v_sv:.0f} V" if v_sv else "")
        ax.annotate(lbl_su, (xu, yu), xytext=(9, -16), textcoords="offset points",
                    fontsize=7, color="darkred",
                    bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="red", alpha=0.85))
        ax.annotate(lbl_sv, (xv, yv), xytext=(9, 7), textcoords="offset points",
                    fontsize=7, color="darkred",
                    bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="red", alpha=0.85))

        # OPEN label at midpoint
        mx, my = (xu + xv) / 2, (yu + yv) / 2
        ax.text(mx, my, f"[OPT] OPEN (#{opt_idx})", fontsize=8, ha="center", va="center",
                color="red", fontweight="bold", zorder=8,
                bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="red", alpha=0.92))

        # Tie line
        xna, yna = net_a.node_coords[opt["tie_node_a"]]
        xnb, ynb = net_b.node_coords[opt["tie_node_b"]]
        ax.plot([xna, xnb], [yna, ynb], color="#00B050", linewidth=3, linestyle="--", zorder=6)
        ax.scatter([xna, xnb], [yna, ynb], color="#00B050", s=160, zorder=7, marker="D")

        # Tie node labels with voltage
        v_na = volt_a.get(opt["tie_node_a"]); v_nb = volt_b.get(opt["tie_node_b"])
        lbl_na = f"TIE_A  nid={opt['tie_node_a']}" + (f"\n{v_na:.0f} V" if v_na else "")
        lbl_nb = f"TIE_B  nid={opt['tie_node_b']}" + (f"\n{v_nb:.0f} V" if v_nb else "")
        ax.annotate(lbl_na, (xna, yna), xytext=(-10, 14), textcoords="offset points",
                    fontsize=7, color="#005C2E",
                    bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="#00B050", alpha=0.85))
        ax.annotate(lbl_nb, (xnb, ynb), xytext=(9, 14), textcoords="offset points",
                    fontsize=7, color="#005C2E",
                    bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="#00B050", alpha=0.85))

        # Summary annotation near tie midpoint
        ax.annotate(
            f"Tie {opt['tie_distance_m']:.1f} m  AW {opt.get('tie_conductor_size', 50)}mm2\n"
            f"{opt['subtree_kw']:.1f} kW  A:{opt['a_loading_after']:.1f}%  "
            f"B:{opt['b_loading_after']:.1f}%  Vmin:{opt['min_v']:.0f}V",
            xy=((xna + xnb) / 2, (yna + ynb) / 2),
            xytext=(14, -22), textcoords="offset points",
            fontsize=8, color="#00B050",
            bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="#00B050", alpha=0.88),
        )

    # --- Colorbar ---
    if has_volt:
        sm = mcm.ScalarMappable(cmap=cmap, norm=norm)
        sm.set_array([])
        cbar = fig.colorbar(sm, ax=ax, shrink=0.62, pad=0.02)
        cbar.set_label("Node Voltage (V)", fontsize=9)
        cbar.ax.axhline(y=200, color="red",      linewidth=1.5, linestyle="--")
        cbar.ax.axhline(y=220, color="darkorange", linewidth=1.0, linestyle=":")
        y_200 = (200 - 150) / (240 - 150)
        y_220 = (220 - 150) / (240 - 150)
        cbar.ax.text(1.08, y_200, "200V", transform=cbar.ax.transAxes,
                     fontsize=7, color="red", va="center")
        cbar.ax.text(1.08, y_220, "220V", transform=cbar.ax.transAxes,
                     fontsize=7, color="darkorange", va="center")

    # --- Legend ---
    legend_items = [
        Line2D([0], [0], color=DEF_A, lw=2,  label=f"TR_A  ({net_a.facilityid})"),
        Line2D([0], [0], color=DEF_B, lw=2,  label=f"TR_B  ({net_b.facilityid})"),
        Line2D([0], [0], color="red",    lw=2.5, linestyle="--", label="Switch [OPT] open"),
        Line2D([0], [0], color="#FF9999",lw=1.5, linestyle=":",  label="Switch [other] open"),
        Line2D([0], [0], color="#00B050",lw=2.5, linestyle="--", label="New tie line (close)"),
        Line2D([0], [0], marker="*", color="gray", markersize=12, lw=0, label="Transformer"),
    ]
    if has_volt:
        legend_items.append(
            Line2D([0], [0], marker="o", color="none", markerfacecolor="none",
                   markeredgecolor="red", markersize=9, markeredgewidth=1.5, lw=0,
                   label="Voltage < 200V"))
    ax.legend(handles=legend_items, loc="upper right", fontsize=8, framealpha=0.9)

    base_txt = ""
    if results:
        base_txt = (f"\nBaseline: A={results[0]['a_loading_before']:.1f}%  "
                    f"B={results[0]['b_loading_before']:.1f}%")
    ax.set_title(
        f"Load Transfer Topology  |  {net_a.facilityid} (A) -> {net_b.facilityid} (B)"
        + base_txt,
        fontsize=11,
    )
    ax.set_xlabel("Easting (UTM47N, m)")
    ax.set_ylabel("Northing (UTM47N, m)")
    ax.ticklabel_format(style="plain", axis="both")
    ax.grid(True, alpha=0.25)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"[Map]   บันทึก: {out_path}")


# =====================================================================
# Interactive HTML map (Plotly)
# =====================================================================

def draw_interactive_map(
    net_a: NetworkGraph,
    net_b: NetworkGraph,
    results: List[dict],
    out_path: str,
    volt_a:  Dict[int, float]            = None,
    volt_b:  Dict[int, float]            = None,
    phase_a: Dict[int, Dict[int, float]] = None,
    phase_b: Dict[int, Dict[int, float]] = None,
    raw_a: dict = None,
    raw_b: dict = None,
    scenario_voltages: List[Tuple] = None,
) -> None:
    # scenario_voltages: list of (volt_a, volt_b, phase_a, phase_b) per top-5 scenario
    # If provided, replaces single-scenario volt_a/volt_b for per-scenario coloring
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    volt_a   = volt_a  or {}
    volt_b   = volt_b  or {}
    phase_a  = phase_a or {}
    phase_b  = phase_b or {}
    has_volt = bool(volt_a or volt_b)
    vmin, vmax = 150, 240
    cscale = [[0.0, "#d73027"], [0.33, "#fee08b"], [0.66, "#91cf60"], [1.0, "#1a9850"]]

    def _phase_hover(phases: Dict[int, float]) -> str:
        if not phases:
            return ""
        lines = [f"V{PHASE_MAP.get(n, str(n))} = {v:.1f} V" for n, v in sorted(phases.items())]
        return "<br>" + "<br>".join(lines)

    fig = make_subplots(
        rows=2, cols=1,
        row_heights=[0.72, 0.28],
        specs=[[{"type": "scatter"}], [{"type": "table"}]],
        vertical_spacing=0.04,
        subplot_titles=["", ""],
    )

    # ── Edges ───────────────────────────────────────────────────────────
    def _add_edges(net, color, grp, name, width=1.8, opacity=0.65):
        xs, ys = [], []
        seen = set()
        for feat_idx, (u, v) in net.lc_feat_edges.items():
            key = (min(u, v), max(u, v))
            if key in seen:
                continue
            seen.add(key)
            if u not in net.node_coords or v not in net.node_coords or u == v:
                continue
            xs += [net.node_coords[u][0], net.node_coords[v][0], None]
            ys += [net.node_coords[u][1], net.node_coords[v][1], None]
        fig.add_trace(go.Scatter(
            x=xs, y=ys, mode="lines",
            line=dict(color=color, width=width), opacity=opacity,
            hoverinfo="skip", showlegend=True, name=name, legendgroup=grp,
        ), row=1, col=1)

    _add_edges(net_a, "#4472C4", "a", f"TR_A Lines ({net_a.facilityid})")
    _add_edges(net_b, "#ED7D31", "b", f"TR_B Lines ({net_b.facilityid})")

    # ── Static nodes (flat color, always visible) ──────────────────────
    for net, lbl, color in [
            (net_a, "TR_A", "#4472C4"),
            (net_b, "TR_B", "#ED7D31")]:
        xs2, ys2, txt2, sz2 = [], [], [], []
        for nid in sorted(net.node_coords.keys()):
            x, y = net.node_coords[nid]
            kw = net.node_kw.get(nid, 0.0)
            xs2.append(x); ys2.append(y)
            txt2.append(f"<b>{lbl}  nid={nid}</b><br>Load={kw:.2f} kW<br>({x:.0f},{y:.0f})")
            sz2.append(7 + min(kw * 0.6, 14))
        if xs2:
            fig.add_trace(go.Scatter(
                x=xs2, y=ys2, mode="markers",
                marker=dict(color=color, size=sz2, opacity=0.35, line=dict(width=0)),
                name=f"{lbl} nodes",
                text=txt2, hovertemplate="%{text}<extra></extra>",
            ), row=1, col=1)

    # ── Meter points (colored by simulated voltage when available) ─────
    for net, raw, lbl, vd in [(net_a, raw_a, "TR_A", volt_a), (net_b, raw_b, "TR_B", volt_b)]:
        if not raw:
            continue
        mx, my, mtxt, mv = [], [], [], []
        for feat_idx, nid in net.load_feat_nodes.items():
            feat = raw["features"][feat_idx]
            g = feat.get("geometry", {})
            fx = g.get("x"); fy = g.get("y")
            if fx is None or fy is None:
                continue
            peano    = str(get_attr(feat, "PEANO",    "") or "").strip()
            peameter = str(get_attr(feat, "PEAMETER", "") or "").strip()
            kw       = float(get_attr(feat, "KWP", 0.0) or 0.0)
            label    = peano or peameter or f"nid={nid}"
            v = vd.get(nid)
            mx.append(fx); my.append(fy)
            mv.append(v)
            mtxt.append(
                f"<b>{lbl} Meter {label}</b><br>kW = {kw:.2f}"
                + (f"<br>V = {v:.1f} V" if v is not None else "")
            )
        if mx:
            sim = any(v is not None for v in mv)
            marker = (
                dict(symbol="diamond",
                     color=[v if v is not None else vmin for v in mv],
                     colorscale=cscale, cmin=vmin, cmax=vmax,
                     size=7, opacity=0.85, line=dict(width=0))
                if sim else
                dict(symbol="diamond", color="gray", size=7, opacity=0.5)
            )
            mlabel = [f"{v:.0f}V" if v is not None else "" for v in mv]
            fig.add_trace(go.Scatter(
                x=mx, y=my, mode="markers+text" if sim else "markers",
                marker=marker,
                name=f"{lbl} Meters" if sim else f"{lbl} Meters (no sim)",
                text=mlabel, textposition="top center",
                textfont=dict(size=8, color="#333333"),
                hovertext=mtxt, hovertemplate="%{hovertext}<extra></extra>",
            ), row=1, col=1)

    # ── Transformers ────────────────────────────────────────────────────
    for net, lbl, color in [
            (net_a, f"TR_A ({net_a.facilityid})", "#4472C4"),
            (net_b, f"TR_B ({net_b.facilityid})", "#ED7D31")]:
        x, y = net.node_coords[net.transformer_node]
        fig.add_trace(go.Scatter(
            x=[x], y=[y], mode="markers+text",
            marker=dict(symbol="triangle-up", size=22, color=color,
                        line=dict(color="black", width=1.5)),
            text=[lbl], textposition="top right",
            textfont=dict(size=10, color=color),
            name=lbl,
            hovertemplate=f"<b>{lbl}</b><br>({x:.0f}, {y:.0f})<extra></extra>",
        ), row=1, col=1)

    # ── Top-5 feasible scenarios: switch + tie (selectable) ──────────────
    opt = next((r for r in results if r.get("optimal")), None)
    feasible_top = [r for r in results if r.get("feasible")][:5]
    SCEN_COLORS  = ["#00B050", "#0070C0", "#7030A0", "#FF6600", "#A0522D"]

    scen_trace_start = len(fig.data)
    # Traces per scenario: voltage_nodes, low_v_rings, switch, tie  (4 each)
    TRACES_PER_SCEN = 5

    for i, r in enumerate(feasible_top):
        color = SCEN_COLORS[i]
        rank  = next(j + 1 for j, x in enumerate(results) if x is r)
        tag   = " ★OPT" if r.get("optimal") else ""
        grp   = f"scen{i + 1}"
        lw    = 5 if r.get("optimal") else 2.5

        # voltage data for this scenario
        sv_data = scenario_voltages[i] if (scenario_voltages and i < len(scenario_voltages)) else None
        va_i  = sv_data[0] if sv_data else {}
        vb_i  = sv_data[1] if sv_data else {}
        pa_i  = sv_data[2] if sv_data else {}
        pb_i  = sv_data[3] if sv_data else {}

        # ── Voltage-colored nodes (per scenario) ──────────────────────
        all_x, all_y, all_v, all_txt, all_sz = [], [], [], [], []
        _vmin_x = _vmin_y = _vmin_v = _vmin_nid = _vmin_lbl = None
        for net, vd, pd, lbl in [
                (net_a, va_i, pa_i, "TR_A"),
                (net_b, vb_i, pb_i, "TR_B")]:
            for nid in sorted(net.node_coords.keys()):
                v = vd.get(nid)
                if v is None:
                    continue
                x, y = net.node_coords[nid]
                kw = net.node_kw.get(nid, 0.0)
                ph = pd.get(nid, {})
                all_x.append(x); all_y.append(y); all_v.append(v)
                all_txt.append(
                    f"<b>#{rank}{tag}  {lbl}  nid={nid}</b>"
                    + _phase_hover(ph)
                    + (f"<br><b>Vmin = {v:.1f} V</b>" if len(ph) > 1 else f"<br>V = {v:.1f} V")
                    + f"<br>Load = {kw:.2f} kW"
                    + f"<br>({x:.0f}, {y:.0f})"
                )
                all_sz.append(8 + min(kw * 0.6, 14))
                if _vmin_v is None or v < _vmin_v:
                    _vmin_x, _vmin_y, _vmin_v, _vmin_nid, _vmin_lbl = x, y, v, nid, lbl

        show_colorbar = (i == 0)
        fig.add_trace(go.Scatter(
            x=all_x, y=all_y, mode="markers",
            marker=dict(
                color=all_v if all_v else [0],
                colorscale=cscale, cmin=vmin, cmax=vmax,
                size=all_sz if all_sz else [8],
                opacity=0.95, line=dict(width=0),
                colorbar=dict(
                    title=dict(text="Voltage (V)", side="right"),
                    tickvals=list(range(150, 241, 10)),
                    thickness=14, len=0.50, x=1.01, outlinewidth=0.5,
                ) if show_colorbar else None,
                showscale=show_colorbar,
            ),
            name=f"#{rank}{tag} Voltages",
            legendgroup=grp,
            text=all_txt,
            hovertemplate="%{text}<extra></extra>",
        ), row=1, col=1)

        # ── Low-voltage rings < 200 V (per scenario) ─────────────────
        lx, ly, lt = [], [], []
        for net, vd, pd, lbl in [
                (net_a, va_i, pa_i, "TR_A"),
                (net_b, vb_i, pb_i, "TR_B")]:
            for nid, v in vd.items():
                if v < 200 and nid in net.node_coords:
                    x, y = net.node_coords[nid]
                    ph = pd.get(nid, {})
                    lx.append(x); ly.append(y)
                    lt.append(f"<b>#{rank} {lbl} nid={nid}  Vmin={v:.1f}V</b>" + _phase_hover(ph))
        fig.add_trace(go.Scatter(
            x=lx, y=ly, mode="markers",
            marker=dict(symbol="circle-open", color="red", size=20, line=dict(width=2.5)),
            name=f"#{rank}{tag} V<200V",
            legendgroup=grp,
            text=lt, hovertemplate="%{text}<extra></extra>",
            visible=True,
        ), row=1, col=1)

        # ── Vmin node highlight (red star) ────────────────────────────
        if _vmin_x is not None:
            _ph = pa_i.get(_vmin_nid, pb_i.get(_vmin_nid, {}))
            _tip = (f"<b>⚠ Vmin Node  #{rank}{tag}</b><br>"
                    f"{_vmin_lbl}  nid={_vmin_nid}<br>"
                    f"<b>V = {_vmin_v:.1f} V</b>"
                    + _phase_hover(_ph)
                    + f"<br>({_vmin_x:.0f}, {_vmin_y:.0f})")
            fig.add_trace(go.Scatter(
                x=[_vmin_x], y=[_vmin_y], mode="markers",
                marker=dict(symbol="star", color="red", size=24,
                            line=dict(width=2, color="#8B0000")),
                name=f"#{rank}{tag} Vmin={_vmin_v:.0f}V",
                legendgroup=grp,
                text=[_tip],
                hovertemplate="%{text}<extra></extra>",
            ), row=1, col=1)
        else:
            fig.add_trace(go.Scatter(x=[], y=[], mode="markers",
                name=f"#{rank}{tag} Vmin", legendgroup=grp,
            ), row=1, col=1)

        # ── Switch trace ───────────────────────────────────────────────
        su, sv = r["switch_edge"]
        xu, yu = net_a.node_coords[su]
        xv, yv = net_a.node_coords[sv]
        ph_su  = pa_i.get(su, {}); ph_sv = pa_i.get(sv, {})

        fig.add_trace(go.Scatter(
            x=[xu, xv], y=[yu, yv], mode="lines+markers",
            line=dict(color=color, width=lw, dash="dash"),
            marker=dict(symbol="x", size=14, color=color, line=dict(width=3)),
            name=f"#{rank}{tag} Switch", legendgroup=grp,
            hovertemplate=(
                f"<b>Scenario #{rank}{tag} — Switch OPEN</b><br>"
                f"<b>sw_u nid={su}</b>" + _phase_hover(ph_su) + "<br>"
                f"<b>sw_v nid={sv}</b>" + _phase_hover(ph_sv)
                + "<extra></extra>"
            ),
        ), row=1, col=1)

        # ── Tie trace ─────────────────────────────────────────────────
        xna, yna = net_a.node_coords[r["tie_node_a"]]
        xnb, ynb = net_b.node_coords[r["tie_node_b"]]
        ph_na = pa_i.get(r["tie_node_a"], {})
        ph_nb = pb_i.get(r["tie_node_b"], {})

        fig.add_trace(go.Scatter(
            x=[xna, xnb], y=[yna, ynb], mode="lines+markers",
            line=dict(color=color, width=lw, dash="dot"),
            marker=dict(symbol="diamond", size=14, color=color),
            name=f"#{rank}{tag} Tie {r['tie_distance_m']:.1f}m", legendgroup=grp,
            hovertemplate=(
                f"<b>Scenario #{rank}{tag} — Tie CLOSE</b><br>"
                f"<b>TIE_A nid={r['tie_node_a']}</b>" + _phase_hover(ph_na) + "<br>"
                f"<b>TIE_B nid={r['tie_node_b']}</b>" + _phase_hover(ph_nb) + "<br>"
                f"Dist: {r['tie_distance_m']:.1f} m  AW {r.get('tie_conductor_size', 50)}mm²<br>"
                f"Transfer: {r['subtree_kw']:.1f} kW<br>"
                f"TR_A: {r['a_loading_after']:.1f}%  TR_B: {r['b_loading_after']:.1f}%<br>"
                f"Vmin: {r.get('min_v', 0):.0f} V"
                "<extra></extra>"
            ),
        ), row=1, col=1)

    scen_trace_end = len(fig.data)

    # ── Summary table ────────────────────────────────────────────────────
    def _scenario_status(r, min_v_thr=200.0, load_lim=80.0):
        if r.get("error"):
            return r["error"][:40], "#ffcccc"
        if not r.get("feasible"):
            reasons = []
            if r.get("a_loading_after", 0) >= load_lim:
                reasons.append(f"TR_A {r['a_loading_after']:.1f}% > {load_lim:.0f}%")
            if r.get("b_loading_after", 0) >= load_lim:
                reasons.append(f"TR_B {r['b_loading_after']:.1f}% > {load_lim:.0f}%")
            if r.get("min_v", 999) < min_v_thr:
                reasons.append(f"Vmin {r['min_v']:.0f}V < {min_v_thr:.0f}V -- ไม่ผ่าน")
            return " | ".join(reasons) or "NG", "#ffcccc"
        if r.get("optimal"):
            return "OPT -- เหมาะสมที่สุด", "#c6efce"
        return "OK", "#e2efda"

    min_v_threshold = opt.get("min_v", 200) if opt else 200   # use actual min for display
    # load limit from optimizer isn't stored in results; use 80 as default display
    tbl_num, tbl_sw, tbl_tie, tbl_dist, tbl_kw = [], [], [], [], []
    tbl_a_bef, tbl_a_aft, tbl_b_bef, tbl_b_aft = [], [], [], []
    tbl_vmin, tbl_csz, tbl_status, tbl_colors = [], [], [], []

    rank = 0
    for r in results:
        rank += 1
        su_t, sv_t = r["switch_edge"]
        status_txt, row_color = _scenario_status(r)
        tbl_num.append(str(rank))
        tbl_sw.append(f"{su_t} → {sv_t}")
        tbl_tie.append(f"{r['tie_node_a']} ↔ {r['tie_node_b']}")
        tbl_dist.append(f"{r['tie_distance_m']:.1f}")
        tbl_kw.append(f"{r['subtree_kw']:.1f}")
        tbl_a_bef.append(f"{r.get('a_loading_before',0):.1f}%")
        tbl_a_aft.append(f"{r.get('a_loading_after', r.get('a_loading_before',0)):.1f}%"
                         if not r.get("error") else "—")
        tbl_b_bef.append(f"{r.get('b_loading_before',0):.1f}%")
        tbl_b_aft.append(f"{r.get('b_loading_after', r.get('b_loading_before',0)):.1f}%"
                         if not r.get("error") else "—")
        tbl_vmin.append(f"{r['min_v']:.0f} V" if "min_v" in r else "—")
        tbl_csz.append(f"AW {r.get('tie_conductor_size', 50)}mm²" if r.get("feasible") else "—")
        tbl_status.append(status_txt)
        tbl_colors.append(row_color)

    hdr_color  = "#1F497D"
    fig.add_trace(go.Table(
        header=dict(
            values=["#", "Switch (u→v)", "Tie (A↔B)", "Dist(m)",
                    "kW xfr", "TR_A bef", "TR_A aft", "TR_B bef", "TR_B aft",
                    "Vmin", "Conductor", "สรุปผล"],
            fill_color=hdr_color,
            font=dict(color="white", size=11),
            align="center", height=28,
        ),
        cells=dict(
            values=[tbl_num, tbl_sw, tbl_tie, tbl_dist, tbl_kw,
                    tbl_a_bef, tbl_a_aft, tbl_b_bef, tbl_b_aft,
                    tbl_vmin, tbl_csz, tbl_status],
            fill_color=[tbl_colors] * 12,
            font=dict(size=11),
            align=["center"] * 10 + ["left"],
            height=24,
        ),
    ), row=2, col=1)

    # ── Layout + Scenario selector buttons ───────────────────────────────
    base_a = results[0].get("a_loading_before", 0) if results else 0
    base_b = results[0].get("b_loading_before", 0) if results else 0

    # Build visibility masks for updatemenus
    # traces: [base...][scen_switch, scen_tie, ...][table]
    n_base   = scen_trace_start
    n_scen   = len(feasible_top)
    n_after  = len(fig.data) - scen_trace_end   # table trace(s)

    def _vis_mask(sel):
        vis = [True] * n_base
        for i in range(n_scen):
            show = (sel is None or sel == i)
            vis += [show] * TRACES_PER_SCEN
        vis += [True] * n_after
        return vis

    scen_buttons = [dict(
        label="แสดงทั้งหมด",
        method="update",
        args=[{"visible": _vis_mask(None)}],
    )]
    for i, r in enumerate(feasible_top):
        rank = next(j + 1 for j, x in enumerate(results) if x is r)
        tag  = "★ " if r.get("optimal") else ""
        lbl  = (
            f"{tag}#{rank}  "
            f"A:{r['a_loading_after']:.0f}%→{r['b_loading_after']:.0f}%  "
            f"Vmin:{r.get('min_v', 0):.0f}V  "
            f"Tie:{r['tie_distance_m']:.0f}m"
        )
        scen_buttons.append(dict(
            label=lbl,
            method="update",
            args=[{"visible": _vis_mask(i)}],
        ))

    fig.update_layout(
        title=dict(
            text=(
                f"Load Transfer Topology  |  {net_a.facilityid} (A) → {net_b.facilityid} (B)  |  "
                f"Baseline: A={base_a:.1f}%  B={base_b:.1f}%"
                f"<br><sup>Scroll to zoom  |  Drag to pan  |  Click legend to toggle layers  |  "
                f"Hover = แรงดันรายเฟส</sup>"
            ),
            font=dict(size=13),
        ),
        hovermode="closest",
        plot_bgcolor="white",
        paper_bgcolor="white",
        height=1050,
        legend=dict(
            x=1.07, y=0.99,
            bgcolor="rgba(255,255,255,0.92)",
            bordercolor="#cccccc", borderwidth=1,
            font=dict(size=11),
        ),
        margin=dict(r=230, t=90, l=75, b=30),
        updatemenus=[dict(
            type="buttons",
            direction="down",
            x=1.16, y=0.58,
            xanchor="left",
            showactive=True,
            bgcolor="white",
            bordercolor="#cccccc",
            font=dict(size=11),
            buttons=scen_buttons,
        )],
        annotations=[dict(
            text="<b>เลือก Scenario</b>",
            x=1.16, y=0.61,
            xref="paper", yref="paper",
            xanchor="left", showarrow=False,
            font=dict(size=11, color="#333333"),
        )],
    )
    fig.update_xaxes(
        title_text="Easting (UTM47N, m)", tickformat="d",
        showgrid=True, gridcolor="#eeeeee", zeroline=False,
        row=1, col=1,
    )
    fig.update_yaxes(
        title_text="Northing (UTM47N, m)", tickformat="d",
        showgrid=True, gridcolor="#eeeeee", zeroline=False,
        scaleanchor="x", scaleratio=1,
        row=1, col=1,
    )

    fig.write_html(out_path, include_plotlyjs="cdn")
    print(f"[Map]   Interactive: {out_path}")


# =====================================================================
# CLI
# =====================================================================

def main() -> None:
    ap = argparse.ArgumentParser(
        description="LV Transformer Neighborhood — Load Transfer Optimizer",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    ap.add_argument("fac_a", help="FACILITYID หม้อแปลงที่โหลดเกิน  (e.g. 04-123456)")
    ap.add_argument("fac_b", help="FACILITYID หม้อแปลงที่รับโหลดได้ (e.g. 04-234567)")
    ap.add_argument("--max-tie-dist", type=float, default=100.0, metavar="M",
                    help="ระยะสูงสุดของสาย tie (m)")
    ap.add_argument("--min-voltage",  type=float, default=210.0,  metavar="V",
                    help="แรงดันต่ำสุดที่ยอมรับ (V)")
    ap.add_argument("--load-limit",   type=float, default=80.0,   metavar="PCT",
                    help="โหลดสูงสุดที่ยอมรับ (%)")
    ap.add_argument("--json-dir",     default=OUTPUT_DIR,
                    help="folder ที่เก็บไฟล์ JSON")
    ap.add_argument("--out-dir",      default=r"D:\TRneighborhood\output",
                    help="folder สำหรับบันทึก Excel/PNG")
    ap.add_argument("--region",       default=None,
                    help="เขต GIS ของ กฟภ. ที่หม้อแปลงทั้งคู่อยู่ (เช่น NE1, C2, Z)")
    args = ap.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fa_s = re.sub(r"[^A-Za-z0-9]", "_", args.fac_a)
    fb_s = re.sub(r"[^A-Za-z0-9]", "_", args.fac_b)
    stem = out_dir / f"transfer_{fa_s}_{fb_s}_{ts}"

    optimizer = TransferOptimizer(
        fac_a=args.fac_a,
        fac_b=args.fac_b,
        max_tie_dist=args.max_tie_dist,
        load_limit_pct=args.load_limit,
        min_voltage_v=args.min_voltage,
        json_dir=args.json_dir,
        region=args.region,
    )
    results = optimizer.run()

    if not results:
        print("\nไม่มีผลลัพธ์ที่จะแสดง")
        sys.exit(0)

    print_results_table(results, optimizer.net_a, optimizer.net_b)
    save_excel_report(results, optimizer.net_a, optimizer.net_b, str(stem) + ".xlsx")

    # Collect per-node voltages for voltage-coloured map
    volt_a_map:  Dict[int, float] = {}
    volt_b_map:  Dict[int, float] = {}
    phase_a_map: Dict[int, Dict[int, float]] = {}
    phase_b_map: Dict[int, Dict[int, float]] = {}
    raw_a_map: dict = None
    raw_b_map: dict = None
    opt = next((r for r in results if r.get("optimal")), None)
    if opt:
        print("\n[Voltage Map] รัน simulation สำหรับ voltage coloring...")
        with open(optimizer._json_a, encoding="utf-8") as fh:
            raw_a_map = json.load(fh)
        with open(optimizer._json_b, encoding="utf-8") as fh:
            raw_b_map = json.load(fh)
        volt_a_map, volt_b_map, phase_a_map, phase_b_map = collect_optimal_node_voltages(
            optimizer.net_a, optimizer.net_b, raw_a_map, raw_b_map, opt,
            snap_tol_a=optimizer.net_a.snap_tol,
            snap_tol_b=optimizer.net_b.snap_tol)
        print(f"  TR_A: {len(volt_a_map)} nodes  TR_B: {len(volt_b_map)} nodes")

    draw_topology_map(optimizer.net_a, optimizer.net_b, results, str(stem) + ".png",
                      volt_a=volt_a_map, volt_b=volt_b_map)
    draw_interactive_map(optimizer.net_a, optimizer.net_b, results, str(stem) + ".html",
                         volt_a=volt_a_map, volt_b=volt_b_map,
                         phase_a=phase_a_map, phase_b=phase_b_map,
                         raw_a=raw_a_map, raw_b=raw_b_map)
    print(f"\nเสร็จสิ้น — ผลลัพธ์บันทึกที่ {out_dir}")


if __name__ == "__main__":
    main()
