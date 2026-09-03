"""
PhaseOptimizer.py
-----------------
วิเคราะห์ระบบจำหน่ายแรงต่ำ (LV) ของหม้อแปลงตัวเดียว
และแนะนำการปรับปรุง 3 ขั้นตอน:
  1. ย้ายเฟสมิเตอร์   (Phase Transfer)  — ย้ายให้สอดคล้องกับสายที่มีเฟสรองรับ
  2. เพิ่มขนาดสาย     (Conductor Upgrade) — 50 → 95 mm²
  3. เพิ่มเฟสสาย      (Phase Addition)   — single/dual-phase → 3-phase พร้อมกระจายโหลด

Pipeline:
  1. โหลด/ดึง JSON network จาก FacilityID
  2. รัน OpenDSS baseline
  3. ตรวจแรงดันตก / phase imbalance
  4. ถ้าพบปัญหา → ลองย้ายเฟสมิเตอร์ (greedy)
  5. ถ้ายังไม่ผ่าน → simulate เพิ่มขนาดสาย
  6. ถ้ายังไม่ผ่าน → simulate เพิ่มเฟสสาย

Usage:
    python PhaseOptimizer.py <FACILITYID> [options]

    --min-voltage   float  แรงดันต่ำสุด (V)          [default 200]
    --max-imbalance float  phase imbalance สูงสุด (%) [default 25]
    --json-dir      path   folder JSON               [default D:\\testpy\\jsonfile]
    --out-dir       path   folder ผลลัพธ์            [default D:\\TRneighborhood\\output]
"""

import os, sys, json, math, re, argparse, tempfile, shutil, contextlib, io, copy, warnings, time, threading
from pathlib import Path
from datetime import datetime
from collections import defaultdict, deque
from dataclasses import dataclass, field
from typing import Dict, FrozenSet, List, Optional, Set, Tuple

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# This file lives in feature_PhaseOptimizer/, but the modules it depends on
# live in feature_shareload/ (Runopendss_All05082026.py, TransferOptimizer-
# 072026.py) and the project root (InputJsonApi.py). Put both on sys.path so
# imports resolve regardless of the caller's cwd.
_THIS_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = str(_THIS_DIR.parent)
_SHARELOAD_DIR = str(_THIS_DIR.parent / "feature_shareload")
for _p in (_PROJECT_ROOT, _SHARELOAD_DIR):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import numpy as np
from scipy.spatial import cKDTree
import networkx as nx
import openpyxl
from openpyxl.styles import PatternFill, Font
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.cm as mcm
import matplotlib.colors as mcolors
from matplotlib.lines import Line2D
import matplotlib.font_manager as _fm

# ตั้งฟอนต์ไทย — ลองตามลำดับความนิยม
_THAI_FONT_PREF = ["TH SarabunPSK", "TH Sarabun New", "Tahoma", "Leelawadee UI", "Leelawadee"]
_available_fonts = {f.name for f in _fm.fontManager.ttflist}
_thai_font = next((f for f in _THAI_FONT_PREF if f in _available_fonts), None)
if _thai_font:
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = [_thai_font, "DejaVu Sans", "Arial"]

# ปิด warning "Glyph X missing from font" — matplotlib ยัง fallback ไปฟอนต์ถัดไปได้เอง
warnings.filterwarnings("ignore", message="Glyph .* missing from font", category=UserWarning)

from Runopendss_All05082026 import (
    get_attr, has_point, has_paths, point_xy, endpoints_from_paths,
    cluster_points, convert_json_to_dss_ordered, solve_with_opendss, build_bfs_order,
    tx_secondary_power_by_node, AW_IMP,
)

# TransferOptimizer-072026.py's filename has a hyphen, so it can't be reached
# with a plain `import` statement — load it by file path (same trick used in
# feature_shareload/run_web.py). Using this file (not the older
# TransferOptimizer.py) keeps PhaseOptimizer on the same NetworkGraph/region-
# aware-fetch/multi-phase-transformer modeling as the shareload feature.
import importlib.util as _ilu
_spec = _ilu.spec_from_file_location(
    "TransferOptimizer_072026", _THIS_DIR.parent / "feature_shareload" / "TransferOptimizer-072026.py"
)
_topt_mod = _ilu.module_from_spec(_spec)
_spec.loader.exec_module(_topt_mod)

NetworkGraph                 = _topt_mod.NetworkGraph
_collect_dss_bus_phase_voltages = _topt_mod._collect_dss_bus_phase_voltages
query_transformer_loading    = _topt_mod.query_transformer_loading
import opendssdirect as odss

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
JSON_DIR    = r"D:\testpy\jsonfile"
DEFAULT_OUT = r"D:\TRneighborhood\output_phase"

PHASE_MAP   = {1: "A", 2: "B", 3: "C"}

_dss_lock = threading.Lock()       # DSS engine เป็น singleton — ต้องใช้ lock เมื่อมีหลาย thread
_AW_SIZES = sorted(AW_IMP.keys())  # [50, 95] — ขนาดสายที่รองรับใน impedance table
PD_TO_NODE  = {4: 1, 2: 2, 1: 3}        # single-phase PHASEDESIGNATION → DSS node#
NODE_TO_PD  = {1: 4, 2: 2, 3: 1}
PD_TO_PHASE = {4: "A", 2: "B", 1: "C"}
SINGLE_PDS  = frozenset({1, 2, 4})       # single-phase PHASEDESIGNATION values

_PD_LABEL = {7: "ABC", 6: "AB", 5: "CA", 4: "A", 3: "BC", 2: "B", 1: "C"}


# ─────────────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class SimResult:
    converged: bool
    node_phase_v: Dict[int, Dict[int, float]]  # node_id → {phase_node: V}
    min_v: float
    tr_loading_pct: float
    low_v_nodes: List[int]                     # nodes where min_phase_V < threshold
    phase_kw: Dict[int, float] = field(default_factory=dict)  # {1:kWA, 2:kWB, 3:kWC}
    phase_imbalance_pct: float = 0.0
    error: str = ""


@dataclass
class MeterInfo:
    feat_idx: int
    node_id: int
    peano: str
    current_pd: int    # PHASEDESIGNATION in JSON (1,2,4 for single-phase)
    conductor_pd: int  # phases physically on connecting conductor
    kw: float

    @property
    def is_single_phase(self) -> bool:
        return self.current_pd in SINGLE_PDS

    @property
    def movable_pds(self) -> List[int]:
        """Other single-phase PDs available on the conductor."""
        other = self.conductor_pd & ~self.current_pd
        return [pd for pd in (4, 2, 1) if other & pd]


@dataclass
class PhaseMove:
    meter: MeterInfo
    from_pd: int
    to_pd: int
    delta_min_v: float
    delta_imbalance: float


@dataclass
class UpgradeResult:
    edge: Tuple[int, int]
    feat_idx: int
    from_size: int
    to_size: int
    n_affected: int
    result: SimResult
    delta_min_v: float


@dataclass
class PhaseAddResult:
    feat_idx: int
    edge: Tuple[int, int]          # target segment (u→v) ที่ low-V อยู่
    from_pd: int
    to_pd: int
    n_affected: int
    meter_moves: List[Tuple[int, int]]          # (feat_idx, new_pd) เฉพาะใน subtree(v)
    upgraded_edges: List[Tuple[int, int, int]]  # (feat_idx, u, v) ทุก edge บน path ที่ต้องอัปเกรด
    result: SimResult
    delta_min_v: float


# ─────────────────────────────────────────────────────────────────────────────
# Helper functions
# ─────────────────────────────────────────────────────────────────────────────

def _phase_imbalance_pct(phases: Dict[int, float]) -> float:
    if len(phases) < 2:
        return 0.0
    vals = list(phases.values())
    mean_v = sum(vals) / len(vals)
    return max(abs(v - mean_v) for v in vals) / mean_v * 100.0 if mean_v else 0.0


def _conductor_pd_at_node(net: NetworkGraph, raw: dict, node_id: int) -> int:
    """Union of PHASEDESIGNATION bits for all LC lines touching node_id."""
    result = 0
    for feat_idx, (u, v) in net.lc_feat_edges.items():
        if u == node_id or v == node_id:
            pd = int(get_attr(raw["features"][feat_idx], "PHASEDESIGNATION", 7) or 7)
            result |= pd
    return result if result else 7


def _build_bus_to_node(net: NetworkGraph) -> Dict[str, int]:
    """bus_name.lower() → node_id mapping from BFS order of original network."""
    lc_edges = list(net.lc_feat_edges.values())
    new_bus_map, _, _, _ = build_bfs_order(
        len(net.node_coords), lc_edges, net.transformer_node)
    return {b.lower(): nid for nid, b in new_bus_map.items()}


def _simulate(
    json_path: str,
    bus_to_node: Dict[str, int],
    rated_kva: float,
    min_v_thr: float,
    snap_tol: float = 2.0,
) -> SimResult:
    """Run OpenDSS on json_path; return SimResult. Topology must match bus_to_node."""
    tmp = tempfile.mkdtemp(prefix="popt_")
    try:
        dss_path = str(Path(tmp) / "sim.dss")
        with _dss_lock:
            with contextlib.redirect_stdout(io.StringIO()):
                dss_file, _ = convert_json_to_dss_ordered(json_path, dss_path, snap_tol=snap_tol)
                solve_with_opendss(dss_file)

            if not odss.Solution.Converged():
                return SimResult(converged=False, node_phase_v={}, min_v=0.0,
                                 tr_loading_pct=0.0, low_v_nodes=[], error="not converged")

            dss_vmap = _collect_dss_bus_phase_voltages()

            # Transformer per-phase kW (secondary terminal). Single/two-phase
            # transformers get split into multiple DSS Transformer objects
            # (one per phase) by convert_json_to_dss_ordered — reading only
            # tx_names[0] would see just one phase's power and (with fewer
            # than 2 phase_kw entries) always report 0% imbalance, silently
            # hiding the exact condition this tool exists to detect. Sum
            # tx_secondary_power_by_node() across every transformer object
            # instead (same aggregation as Runopendss's report_transformer_pattern).
            phase_kw: Dict[int, float] = {}
            phase_imbal = 0.0
            try:
                tx_names = odss.Transformers.AllNames()
                for _name in tx_names:
                    for node, (p, _q) in tx_secondary_power_by_node(f"Transformer.{_name}").items():
                        phase_kw[node] = phase_kw.get(node, 0.0) + abs(p)
                if len(phase_kw) >= 2:
                    vals = list(phase_kw.values())
                    mv   = sum(vals) / len(vals)
                    phase_imbal = (max(abs(v - mv) for v in vals) / mv * 100.0
                                   if mv else 0.0)
            except Exception:
                pass

            tr_loading = query_transformer_loading(rated_kva)

        # Map bus names → node IDs outside lock (pure Python, no DSS calls)
        node_phase_v: Dict[int, Dict[int, float]] = {}
        for bus, phases in dss_vmap.items():
            nid = bus_to_node.get(bus.lower())
            if nid is not None:
                node_phase_v[nid] = phases

        min_v_global = min(
            (min(ph.values()) for ph in node_phase_v.values() if ph),
            default=0.0,
        )
        low_v_nodes = [
            nid for nid, ph in node_phase_v.items()
            if ph and min(ph.values()) < min_v_thr
        ]

        return SimResult(
            converged=True,
            node_phase_v=node_phase_v,
            min_v=min_v_global,
            tr_loading_pct=tr_loading,
            low_v_nodes=low_v_nodes,
            phase_kw=phase_kw,
            phase_imbalance_pct=phase_imbal,
        )
    except Exception as exc:
        return SimResult(converged=False, node_phase_v={}, min_v=0.0,
                         tr_loading_pct=0.0, low_v_nodes=[], error=str(exc))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _sim_modified_json(
    modified_raw: dict,
    bus_to_node: Dict[str, int],
    rated_kva: float,
    min_v_thr: float,
    snap_tol: float,
) -> SimResult:
    """Write modified_raw to temp file and simulate."""
    tmp_json = tempfile.NamedTemporaryFile(
        suffix=".json", delete=False, mode="w", encoding="utf-8")
    json.dump(modified_raw, tmp_json, ensure_ascii=False)
    tmp_json.close()
    try:
        return _simulate(tmp_json.name, bus_to_node, rated_kva, min_v_thr, snap_tol)
    finally:
        os.unlink(tmp_json.name)


_LETTER_TO_PD = {"A": 4, "B": 2, "C": 1}


def _meter_current_pd(feat: dict) -> int:
    """
    Determine effective single-phase PHASEDESIGNATION for a meter feature.
    DSS generator uses PHASE attribute (A/B/C) before PHASEDESIGNATION,
    so we mirror the same priority to get the correct current phase.
    """
    letter = str(get_attr(feat, "PHASE", "") or "").strip().upper()
    if letter in _LETTER_TO_PD:
        return _LETTER_TO_PD[letter]
    pd = int(get_attr(feat, "PHASEDESIGNATION", 7) or 7)
    # If not a clean single-phase PD, infer from first available phase bit
    if pd not in SINGLE_PDS:
        for bit in (4, 2, 1):
            if pd & bit:
                return bit
    return pd


def _apply_meter_phase(raw_features: list, feat_idx: int, new_pd: int) -> None:
    """Update both PHASEDESIGNATION and PHASE on a meter feature in-place."""
    attrs = raw_features[feat_idx].setdefault("attributes", {})
    attrs["PHASEDESIGNATION"] = new_pd
    attrs["PHASE"] = PD_TO_PHASE.get(new_pd, "A")


def _build_meter_inventory(net: NetworkGraph, raw: dict) -> List[MeterInfo]:
    """Collect all meters with their phase and conductor availability."""
    meters: List[MeterInfo] = []
    for feat_idx, node_id in net.load_feat_nodes.items():
        feat = raw["features"][feat_idx]
        peano = str(get_attr(feat, "PEANO", "") or "").strip()
        pd = _meter_current_pd(feat)
        kw = float(get_attr(feat, "KWP", 0.0) or 0.0)
        cond_pd = _conductor_pd_at_node(net, raw, node_id)
        meters.append(MeterInfo(
            feat_idx=feat_idx, node_id=node_id, peano=peano,
            current_pd=pd, conductor_pd=cond_pd, kw=kw,
        ))
    return meters


def _score_result(before: SimResult, after: SimResult) -> float:
    """Score improvement (higher = better). Prioritise V_min, then imbalance."""
    if not after.converged:
        return -999.0
    return (after.min_v - before.min_v) * 10.0 + \
           (before.phase_imbalance_pct - after.phase_imbalance_pct)


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — Phase Transfer Optimizer
# ─────────────────────────────────────────────────────────────────────────────

def optimize_phase_transfer(
    net: NetworkGraph,
    raw: dict,
    baseline: SimResult,
    bus_to_node: Dict[str, int],
    min_v_thr: float,
    max_iters: int = 20,
    max_candidates_per_iter: int = 5,
    snap_tol: float = 2.0,
    max_imbalance_pct: float = 5.0,
    time_limit_s: float = 300.0,
) -> Tuple[dict, List[PhaseMove]]:
    """
    Greedy phase transfer: repeatedly pick the single best meter move
    that most improves V_min / imbalance.

    Only considers meters whose nodes are in or adjacent to low-voltage nodes,
    and meters on the most-loaded phase when imbalance > threshold.

    Returns (modified_raw, list_of_moves).
    """
    current_raw    = copy.deepcopy(raw)
    current_result = baseline
    moves: List[PhaseMove] = []

    all_meters = _build_meter_inventory(net, current_raw)
    movable    = [m for m in all_meters if m.is_single_phase and m.movable_pds]

    print(f"\n[Phase Transfer] มิเตอร์ทั้งหมด={len(all_meters)}  "
          f"single-phase={sum(1 for m in all_meters if m.is_single_phase)}  "
          f"ย้ายได้={len(movable)}")

    if not movable:
        print("  ไม่มีมิเตอร์ที่ย้ายได้ — ข้าม")
        return current_raw, moves

    _t0 = time.time()
    for iteration in range(max_iters):
        if time.time() - _t0 > time_limit_s:
            print(f"  [iter {iteration}] เกิน time limit {time_limit_s:.0f}s — หยุด")
            break
        # Stop if no low-voltage nodes and imbalance acceptable
        has_problem = (bool(current_result.low_v_nodes) or
                       current_result.phase_imbalance_pct > max_imbalance_pct)
        if not has_problem:
            print(f"  [iter {iteration}] ระบบปกติแล้ว — หยุด")
            break

        # Narrow candidates: meters near low-V nodes or on most-loaded phase
        low_v_set = set(current_result.low_v_nodes)
        neighbor_set = set(low_v_set)
        for nid in low_v_set:
            for parent in net.G.predecessors(nid):
                neighbor_set.add(parent)
            for child in net.G.successors(nid):
                neighbor_set.add(child)

        # If imbalance-only problem (no low-V), target meters on heaviest phase
        if not low_v_set and current_result.phase_kw:
            heavy_node = max(current_result.phase_kw, key=current_result.phase_kw.get)
            heavy_pd   = NODE_TO_PD.get(heavy_node, 0)
            focus = [m for m in movable if m.current_pd == heavy_pd]
        else:
            focus = [m for m in movable if m.node_id in neighbor_set]

        if not focus:
            focus = movable  # fallback: try all

        # Sort by kW descending (larger loads have more impact)
        focus.sort(key=lambda m: -m.kw)
        candidates = focus[:max_candidates_per_iter]

        best_score  = 0.02       # minimum improvement threshold
        best_move: Optional[Tuple[MeterInfo, int, SimResult, dict]] = None

        for meter in candidates:
            for new_pd in meter.movable_pds:
                candidate_raw = copy.deepcopy(current_raw)
                _apply_meter_phase(candidate_raw["features"], meter.feat_idx, new_pd)

                result = _sim_modified_json(
                    candidate_raw, bus_to_node,
                    net.rated_kva, min_v_thr, snap_tol)
                score = _score_result(current_result, result)

                if score > best_score:
                    best_score = score
                    best_move  = (meter, new_pd, result, candidate_raw)

        if best_move is None:
            print(f"  [iter {iteration+1}] ไม่มีการย้ายที่ดีกว่า — หยุด")
            break

        meter, new_pd, result, candidate_raw = best_move
        from_pd = meter.current_pd
        dv      = result.min_v - current_result.min_v
        di      = current_result.phase_imbalance_pct - result.phase_imbalance_pct

        moves.append(PhaseMove(meter=meter, from_pd=from_pd, to_pd=new_pd,
                               delta_min_v=dv, delta_imbalance=di))
        print(f"  [iter {iteration+1}] {meter.peano or 'idx='+str(meter.feat_idx):>14}  "
              f"{PD_TO_PHASE.get(from_pd,'?')}→{PD_TO_PHASE.get(new_pd,'?')}  "
              f"ΔV={dv:+.1f}V  Δimbal={di:+.1f}%  Vmin={result.min_v:.1f}V")

        current_result   = result
        current_raw      = candidate_raw
        meter.current_pd = new_pd  # update in-memory

    return current_raw, moves


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — Conductor Upgrade Simulator
# ─────────────────────────────────────────────────────────────────────────────

def simulate_conductor_upgrade(
    net: NetworkGraph,
    raw: dict,
    baseline: SimResult,
    bus_to_node: Dict[str, int],
    min_v_thr: float,
    snap_tol: float = 2.0,
) -> List[UpgradeResult]:
    """
    For each edge that feeds into a low-voltage subtree, try upgrading
    conductor size from 50 → 95 mm².
    Only tests edges where current size < 95; keeps shallowest entry per subtree.
    Returns list sorted by delta_min_v descending.
    """
    results: List[UpgradeResult] = []
    low_v_set = set(baseline.low_v_nodes)

    if not low_v_set:
        print("\n[Conductor Upgrade] ไม่พบโหนดแรงดันต่ำ — ข้าม")
        return []

    # Find edges (u→v) in BFS tree where v's subtree contains low-V nodes
    # Deduplicate: skip subtrees already covered by an ancestor edge
    processed: Set[int] = set()
    candidate_edges: List[Tuple[int, int]] = []

    # BFS order from transformer so we process shallow edges first
    for (u, v) in nx.bfs_edges(net.G, net.transformer_node):
        if v in processed:
            continue
        subtree_v = net.get_subtree((u, v)) | {v}
        if subtree_v & low_v_set:
            candidate_edges.append((u, v))
            processed.update(subtree_v)

    print(f"\n[Conductor Upgrade] ทดสอบ {len(candidate_edges)} สาย (50→95 mm²)")

    for (u, v) in candidate_edges:
        feat_idx = net.find_switch_edge_feature((u, v))
        if feat_idx is None:
            continue
        feat = raw["features"][feat_idx]
        current_size = int(get_attr(feat, "CONDUCTORSIZE", 50) or 50)
        next_sizes = [s for s in _AW_SIZES if s > current_size]
        if not next_sizes:
            print(f"  สาย ({u}→{v}) ขนาด {current_size}mm² — ใหญ่สุดใน AW_IMP แล้ว ข้าม")
            continue
        target_size = next_sizes[0]

        subtree = net.get_subtree((u, v)) | {v}
        n_affected = len(subtree & low_v_set)

        candidate_raw = copy.deepcopy(raw)
        candidate_raw["features"][feat_idx].setdefault("attributes", {})["CONDUCTORSIZE"] = target_size

        result = _sim_modified_json(candidate_raw, bus_to_node,
                                    net.rated_kva, min_v_thr, snap_tol)
        delta_v = result.min_v - baseline.min_v if result.converged else -999.0

        status = (f"Vmin={result.min_v:.1f}V  ΔV={delta_v:+.1f}V"
                  if result.converged else f"[NG] {result.error}")
        print(f"  สาย ({u}→{v}) {current_size}→{target_size}mm²  {n_affected}โหนด LowV  {status}")

        results.append(UpgradeResult(
            edge=(u, v), feat_idx=feat_idx,
            from_size=current_size, to_size=target_size,
            n_affected=n_affected, result=result, delta_min_v=delta_v,
        ))

    results.sort(key=lambda r: -r.delta_min_v)
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Step 3 — Phase Addition Simulator
# ─────────────────────────────────────────────────────────────────────────────

def _path_to_node(net: NetworkGraph, target: int) -> List[Tuple[int, int, int]]:
    """Return [(feat_idx, u, v), ...] for edges on BFS path from transformer_node to target.
    u→v follows BFS tree direction (not necessarily the physical cable direction)."""
    try:
        nodes = nx.shortest_path(net.G, source=net.transformer_node, target=target)
    except (nx.NetworkXNoPath, nx.NodeNotFound):
        return []
    # Bidirectional lookup — lc_feat_edges may store physical cable direction (reversed vs BFS)
    edge_to_feat: Dict[Tuple[int,int], int] = {}
    for fi, (u, v) in net.lc_feat_edges.items():
        edge_to_feat[(u, v)] = fi
        edge_to_feat[(v, u)] = fi
    return [(edge_to_feat[(nodes[i], nodes[i+1])], nodes[i], nodes[i+1])
            for i in range(len(nodes) - 1)
            if (nodes[i], nodes[i+1]) in edge_to_feat]


def simulate_phase_addition(
    net: NetworkGraph,
    raw: dict,
    baseline: SimResult,
    bus_to_node: Dict[str, int],
    min_v_thr: float,
    snap_tol: float = 2.0,
) -> List[PhaseAddResult]:
    """
    For non-3-phase line segments in low-voltage areas:
    1. Upgrade PHASEDESIGNATION → 7 (ABC)
    2. Redistribute meters on that segment evenly across A/B/C by kW
    3. Simulate and measure improvement.

    Returns list sorted by delta_min_v descending.
    """
    results: List[PhaseAddResult] = []
    low_v_set = set(baseline.low_v_nodes)

    if not low_v_set:
        print("\n[Phase Addition] ไม่พบโหนดแรงดันต่ำ — ข้าม")
        return []

    # ── สร้าง undirected graph ของ edges ที่ไม่ใช่ 3 เฟส ──────────────────
    # ต้องอัปเกรดทั้ง connected component พร้อมกัน
    # (ไม่สามารถอัปเกรดกลางสายได้ เพราะจะเกิด A→ABC→A)
    non3_graph: nx.Graph = nx.Graph()
    non3_info: Dict[int, Tuple[int, int, int]] = {}  # feat_idx → (u, v, pd) in BFS direction

    for feat_idx, (pu, pv) in net.lc_feat_edges.items():
        pd = int(get_attr(raw["features"][feat_idx], "PHASEDESIGNATION", 7) or 7)
        if bin(pd).count("1") >= 3:
            continue  # 3-phase อยู่แล้ว ไม่เพิ่ม
        if net.G.has_edge(pu, pv):
            u, v = pu, pv
        elif net.G.has_edge(pv, pu):
            u, v = pv, pu
        else:
            continue
        non3_graph.add_edge(u, v)
        non3_info[feat_idx] = (u, v, pd)

    # ── รวบรวม ALL connected components ของสาย non-3-phase ──────────────────
    all_comps: List[Tuple[Set[int], List]] = []
    for comp_nodes_set in nx.connected_components(non3_graph):
        comp_edges = [(fi, u, v, pd) for fi, (u, v, pd) in non3_info.items()
                      if u in comp_nodes_set and v in comp_nodes_set]
        if comp_edges:
            all_comps.append((comp_nodes_set, comp_edges))

    # ── เลือกเฉพาะ component ที่มี low-V node ────────────────────────────────
    # แต่ละ connected component เริ่มจาก junction กับสาย 3-phase (tab แยก)
    # ไปถึงปลายทาง — entry edge เป็น 3-phase อยู่แล้วโดย definition ของ component
    # ไม่ต้องตรวจ upstream: ถ้าสาย non-3-phase ต่อกันโดยตรงมันจะอยู่ใน component เดียวกัน
    lowv_comp_cis = [ci for ci, (nodes, _) in enumerate(all_comps)
                     if nodes & low_v_set]

    if not lowv_comp_cis:
        print("\n[Phase Addition] ไม่พบ component ที่มีโหนด low-V — ข้าม")
        return []

    # ── Build test cases: เฉพาะ low-V component แต่ละกลุ่ม ─────────────────
    # tuple: (nodes, edges, is_all, [], orig_comp_idx)
    test_cases: List[Tuple[Set[int], List, bool, List[int], int]] = []
    for ci in lowv_comp_cis:
        comp_nodes, comp_edges = all_comps[ci]
        test_cases.append((comp_nodes, comp_edges, False, [], ci))

    # ถ้ามีหลายกลุ่ม → เพิ่มตัวเลือกรวมทุกกลุ่มที่มีปัญหา
    if len(lowv_comp_cis) > 1:
        comb_nodes: Set[int] = set()
        comb_edges: List = []
        for ci in lowv_comp_cis:
            comb_nodes |= all_comps[ci][0]
            comb_edges.extend(all_comps[ci][1])
        test_cases.append((comb_nodes, comb_edges, True, [], -1))

    n_all = len(all_comps)
    n_lowv = len(lowv_comp_cis)
    print(f"\n[Phase Addition] พบ {n_all} กลุ่ม non-3-phase  "
          f"มีปัญหา {n_lowv} กลุ่ม  (ทดสอบ {len(test_cases)} ตัวเลือก)")

    def _run_comp(comp_nodes, comp_edges, label):
        n_affected = len(comp_nodes & low_v_set)
        rep_fi, rep_u, rep_v, rep_pd = comp_edges[0]

        # Redistribute ทุกมิเตอร์ใน component (ทุก node มี A/B/C หลัง upgrade)
        meters_here: List[Tuple[int, float]] = []
        for fi, nid in net.load_feat_nodes.items():
            if nid in comp_nodes:
                m_pd = _meter_current_pd(raw["features"][fi])
                if m_pd in SINGLE_PDS:
                    kw = float(get_attr(raw["features"][fi], "KWP", 0.0) or 0.0)
                    meters_here.append((fi, kw))

        phase_kw_used = {4: 0.0, 2: 0.0, 1: 0.0}
        meter_moves: List[Tuple[int, int]] = []
        for fi, kw in sorted(meters_here, key=lambda x: -x[1]):
            best_pd = min([4, 2, 1], key=lambda p: phase_kw_used[p])
            meter_moves.append((fi, best_pd))
            phase_kw_used[best_pd] += kw

        candidate_raw = copy.deepcopy(raw)
        for fi, eu, ev, _ in comp_edges:
            candidate_raw["features"][fi].setdefault("attributes", {})["PHASEDESIGNATION"] = 7
        for fi, new_m_pd in meter_moves:
            _apply_meter_phase(candidate_raw["features"], fi, new_m_pd)

        result  = _sim_modified_json(candidate_raw, bus_to_node,
                                     net.rated_kva, min_v_thr, snap_tol)
        delta_v = result.min_v - baseline.min_v if result.converged else -999.0
        status  = (f"Vmin={result.min_v:.1f}V  ΔV={delta_v:+.1f}V  "
                   f"ย้าย {len(meter_moves)} มิเตอร์"
                   if result.converged else f"[NG] {result.error}")
        print(f"  {label}  {n_affected}LowV  {status}")
        return PhaseAddResult(
            feat_idx=rep_fi, edge=(rep_u, rep_v),
            from_pd=rep_pd, to_pd=7,
            n_affected=n_affected, meter_moves=meter_moves,
            upgraded_edges=[(fi, eu, ev) for fi, eu, ev, _ in comp_edges],
            result=result, delta_min_v=delta_v,
        )

    for i, (comp_nodes, comp_edges, is_all, dep_cis, orig_ci) in enumerate(test_cases, 1):
        if is_all:
            label = (f"[รวม {len(all_comps)} กลุ่ม] "
                     f"({len(comp_nodes)} nodes, {len(comp_edges)} edges)")
        elif dep_cis:
            dep_str = "+".join(str(d + 1) for d in dep_cis)
            label = (f"[กลุ่ม {orig_ci+1}+ต้นน้ำ{dep_str}] "
                     f"({len(comp_nodes)} nodes, {len(comp_edges)} edges)")
        else:
            label = f"[กลุ่ม {orig_ci+1}] ({len(comp_nodes)} nodes, {len(comp_edges)} edges)"
        results.append(_run_comp(comp_nodes, comp_edges, label))

    results.sort(key=lambda r: -r.delta_min_v)
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Main orchestrator
# ─────────────────────────────────────────────────────────────────────────────

class LVOptimizer:
    """
    Analyse one LV transformer network and produce improvement recommendations.

    Run order:
      1. Fetch/parse JSON  →  2. Baseline DSS  →  3. Phase transfer
      →  4. Conductor upgrade (if still problems)
      →  5. Phase addition   (if still problems)
    """

    def __init__(
        self,
        facilityid: str,
        min_voltage_v: float = 200.0,
        max_imbalance_pct: float = 25.0,
        json_dir: str = JSON_DIR,
        max_sim_time_s: float = 300.0,
        region: Optional[str] = None,
    ) -> None:
        self.facilityid        = facilityid
        self.min_voltage_v     = min_voltage_v
        self.max_imbalance_pct = max_imbalance_pct
        self.json_dir          = json_dir
        self.max_sim_time_s    = max_sim_time_s
        # PEA GIS region (e.g. "NE1", "C2", "Z") the facility lives in — needed
        # to point InputJsonApi at the right regional GIS server before querying;
        # without it, InputJsonApi silently defaults to NE1 and any facility in
        # another region fails with "ไม่พบ TR สำหรับ FACILITYID=...".
        self.region             = region

        self.net: Optional[NetworkGraph]    = None
        self.json_path: Optional[Path]      = None
        self.raw: Optional[dict]            = None
        self.bus_to_node: Dict[str, int]    = {}
        self.snap_tol: float                = 2.0

        self.baseline: Optional[SimResult]  = None
        self.phase_raw: Optional[dict]      = None
        self.phase_result: Optional[SimResult] = None
        self.phase_moves: List[PhaseMove]   = []
        self.upgrade_options: List[UpgradeResult]  = []
        self.applied_upgrade: Optional[UpgradeResult] = None
        self.upgrade_result: Optional[SimResult] = None
        self.phase_moves_cu: List[PhaseMove] = []
        self.phase_result_cu: Optional[SimResult] = None
        self.phase_add_options: List[PhaseAddResult] = []
        self.applied_phase_add: Optional[PhaseAddResult] = None
        self.phase_moves2: List[PhaseMove]  = []        # phase transfer รอบ 2 (หลัง phase addition)
        self.phase_result2: Optional[SimResult] = None
        self.final_raw: Optional[dict]           = None  # final modified JSON after all steps
        self.final_result: Optional[SimResult]   = None  # combined result after all applied steps
        self.error: Optional[str]                = None  # set when run() bails out early — the
                                                           # actual reason (JSON fetch failure,
                                                           # baseline not converging), not just
                                                           # "net/baseline is still None"

    # ------------------------------------------------------------------

    def _ensure_json(self) -> Path:
        fac = self.facilityid
        candidates = [
            Path(self.json_dir) / f"NetworkLV{fac}_with_MV.json",
            Path(self.json_dir) / f"NetworkLV_{fac}_with_MV.json",
            Path(self.json_dir) / f"NetworkLV{fac}.json",
            Path(self.json_dir) / f"NetworkLV_{fac}.json",
        ]
        # ลบ cache เดิมก่อนดึงใหม่เสมอ (แค่ legacy on-disk cache — ไม่ได้อ่านซ้ำ
        # อยู่แล้วเพราะดึงสดจาก API ทุกครั้งด้านล่าง)
        for p in candidates:
            if p.exists():
                p.unlink()
                print(f"  [JSON] ลบ cache เดิม: {p.name}")

        import InputJsonApi as inf
        # Point InputJsonApi at the right regional GIS server first — it
        # defaults to NE1 otherwise, so any facility outside NE1 would fail.
        inf.set_gis_region(self.region)
        print(f"  [JSON] ดึงข้อมูลจาก GIS API FACILITYID: {fac} (region={self.region})...")
        try:
            result = inf.run_once_with_facilityid(fac, project_id=f"phaseopt_{fac}")
        except Exception as exc:
            raise FileNotFoundError(f"ไม่พบ JSON สำหรับ {fac}: {exc}") from exc

        # InputJsonApi writes to pea_no_projects/input/<project_id>/... and
        # returns that exact path — use it directly. Resolve to absolute
        # immediately: OpenDSS's Compile command changes the process cwd as
        # a side effect (see solve_with_opendss), so a relative path cached
        # here would silently break on a later call in the same run.
        out_path = result.get("out_path") if isinstance(result, dict) else None
        if out_path and Path(out_path).exists():
            return Path(out_path).resolve()

        for p in candidates:
            if p.exists():
                return p.resolve()
        raise FileNotFoundError(f"ไม่พบ JSON สำหรับ {fac} แม้หลังดึงข้อมูลจาก API")

    # ------------------------------------------------------------------

    def _has_problem(self, result: SimResult) -> bool:
        return (bool(result.low_v_nodes) or
                result.phase_imbalance_pct > self.max_imbalance_pct)

    def run(self) -> None:
        print(f"\n{'='*60}")
        print(f"LV Phase / Conductor Optimizer")
        print(f"  FACILITYID   : {self.facilityid}")
        print(f"  min_voltage  : {self.min_voltage_v} V")
        print(f"  max_imbalance: {self.max_imbalance_pct}%")
        print(f"{'='*60}")

        # ── Step 1: JSON ─────────────────────────────────────────────────
        print("\n[Step 1] Find TR & LV Network in JSON...")
        try:
            self.json_path = self._ensure_json()
        except FileNotFoundError as exc:
            print(f"  [!] {exc}")
            self.error = str(exc)
            return

        # ── Step 2: Parse network ─────────────────────────────────────────
        print("\n[Step 2] Parse network...")
        self.net = NetworkGraph(str(self.json_path))
        self.snap_tol = self.net.snap_tol
        print(f"  nodes={len(self.net.G.nodes())}  edges={len(self.net.G.edges())}  "
              f"kVA={self.net.rated_kva:.0f}  meters={len(self.net.load_feat_nodes)}")

        with open(self.json_path, encoding="utf-8") as fh:
            self.raw = json.load(fh)

        self.bus_to_node = _build_bus_to_node(self.net)

        # ── Step 3: Baseline simulation ───────────────────────────────────
        print("\n[Step 3] Baseline simulation...")
        self.baseline = _simulate(
            str(self.json_path), self.bus_to_node,
            self.net.rated_kva, self.min_voltage_v, self.snap_tol)
        _print_sim_line("Baseline", self.baseline)

        if not self.baseline.converged:
            print(f"  [!] Baseline ไม่ converge: {self.baseline.error}")
            self.error = f"Baseline ไม่ converge: {self.baseline.error}"
            return

        if not self._has_problem(self.baseline):
            print("\n  [OK] ระบบปกติ — ไม่มีแรงดันตกหรือ phase imbalance เกินเกณฑ์")
            return

        # ── Step 4: Phase transfer ─────────────────────────────────────────
        print("\n[Step 4] ย้ายเฟสมิเตอร์ (Phase Transfer)...")
        self.phase_raw, self.phase_moves = optimize_phase_transfer(
            self.net, self.raw, self.baseline, self.bus_to_node,
            self.min_voltage_v, snap_tol=self.snap_tol,
            max_imbalance_pct=self.max_imbalance_pct, time_limit_s=self.max_sim_time_s)

        # Re-simulate combined result after all moves
        if self.phase_moves:
            self.phase_result = _sim_modified_json(
                self.phase_raw, self.bus_to_node,
                self.net.rated_kva, self.min_voltage_v, self.snap_tol)
            _print_sim_line("After Phase Transfer", self.phase_result)
        else:
            self.phase_result = self.baseline

        # Source for further analysis: use post-phase-transfer state
        analysis_raw    = self.phase_raw or self.raw
        analysis_result = self.phase_result
        self.final_raw  = analysis_raw

        # ── Step 5: Conductor upgrade (if still problems) ─────────────────
        if self._has_problem(analysis_result):
            print("\n[Step 5] เพิ่มขนาดสาย (Conductor Upgrade)...")
            self.upgrade_options = simulate_conductor_upgrade(
                self.net, analysis_raw, analysis_result,
                self.bus_to_node, self.min_voltage_v, self.snap_tol)

        # ── Step 5b: Apply best conductor upgrade if it helps ─────────────
        if self.upgrade_options and self.upgrade_options[0].delta_min_v > 0:
            best_up = self.upgrade_options[0]
            self.applied_upgrade = best_up
            u_up, v_up = best_up.edge
            print(f"\n[Step 5b] ใช้ Conductor Upgrade: edge ({u_up}→{v_up})  "
                  f"{best_up.from_size}→{best_up.to_size}mm²  "
                  f"{best_up.n_affected} โหนด LowV")
            upgrade_raw = copy.deepcopy(analysis_raw)
            upgrade_raw["features"][best_up.feat_idx].setdefault(
                "attributes", {})["CONDUCTORSIZE"] = best_up.to_size

            self.upgrade_result = _sim_modified_json(
                upgrade_raw, self.bus_to_node,
                self.net.rated_kva, self.min_voltage_v, self.snap_tol)
            _print_sim_line("After Conductor Upgrade", self.upgrade_result)

            # ── Step 5c: Phase Transfer หลัง Conductor Upgrade ───────────
            print("\n[Step 5c] Phase Transfer หลัง Conductor Upgrade...")
            upgrade_raw2, self.phase_moves_cu = optimize_phase_transfer(
                self.net, upgrade_raw, self.upgrade_result, self.bus_to_node,
                self.min_voltage_v, snap_tol=self.snap_tol,
                max_imbalance_pct=self.max_imbalance_pct, time_limit_s=self.max_sim_time_s)
            if self.phase_moves_cu:
                self.phase_result_cu = _sim_modified_json(
                    upgrade_raw2, self.bus_to_node,
                    self.net.rated_kva, self.min_voltage_v, self.snap_tol)
                _print_sim_line("After PT (post-Upgrade)", self.phase_result_cu)
                analysis_raw    = upgrade_raw2
                analysis_result = self.phase_result_cu
            else:
                self.phase_result_cu = self.upgrade_result
                analysis_raw    = upgrade_raw
                analysis_result = self.upgrade_result
            self.final_raw = analysis_raw

        # ── Step 6: Phase addition (if still problems) ────────────────────
        if self._has_problem(analysis_result):
            print("\n[Step 6] เพิ่มเฟสสาย (Phase Addition)...")
            self.phase_add_options = simulate_phase_addition(
                self.net, analysis_raw, analysis_result,
                self.bus_to_node, self.min_voltage_v, self.snap_tol)

        # ── Step 7: Apply best phase addition if it actually improves ──────
        if self.phase_add_options and self.phase_add_options[0].delta_min_v > 0:
            best_pa = self.phase_add_options[0]
            self.applied_phase_add = best_pa
            n_up = len(best_pa.upgraded_edges)
            print(f"\n[Step 7] ใช้ Phase Addition: {n_up} edges→ABC  "
                  f"ย้าย {len(best_pa.meter_moves)} มิเตอร์")
            _print_sim_line("After Phase Addition", best_pa.result)

            # ── Step 8: Phase Transfer รอบ 2 หลัง Phase Addition ──────────
            print("\n[Step 8] Phase Transfer รอบ 2 หลัง Phase Addition...")
            combined_raw = copy.deepcopy(analysis_raw)
            for fi, eu, ev in best_pa.upgraded_edges:
                combined_raw["features"][fi].setdefault("attributes", {})["PHASEDESIGNATION"] = 7
            for fi, new_m_pd in best_pa.meter_moves:
                _apply_meter_phase(combined_raw["features"], fi, new_m_pd)

            combined_raw2, self.phase_moves2 = optimize_phase_transfer(
                self.net, combined_raw, best_pa.result, self.bus_to_node,
                self.min_voltage_v, snap_tol=self.snap_tol,
                max_imbalance_pct=self.max_imbalance_pct, time_limit_s=self.max_sim_time_s)

            if self.phase_moves2:
                self.phase_result2 = _sim_modified_json(
                    combined_raw2, self.bus_to_node,
                    self.net.rated_kva, self.min_voltage_v, self.snap_tol)
                _print_sim_line("Combined Final", self.phase_result2)
                self.final_result = self.phase_result2
                self.final_raw    = combined_raw2
            else:
                self.phase_result2 = best_pa.result
                self.final_result  = best_pa.result
                self.final_raw     = combined_raw
                print("  ไม่มีการย้ายเพิ่มเติม")
        else:
            if self.applied_upgrade:
                self.final_result = (self.phase_result_cu
                                     if self.phase_result_cu else self.upgrade_result)
            else:
                self.final_result = self.phase_result if self.phase_result else self.baseline


# ─────────────────────────────────────────────────────────────────────────────
# Console helpers
# ─────────────────────────────────────────────────────────────────────────────

def _print_sim_line(label: str, r: SimResult) -> None:
    if not r.converged:
        print(f"  [{label}] ไม่ converge: {r.error}")
        return
    phase_str = "  ".join(
        f"P{PHASE_MAP.get(n,'?')}={kw:.1f}kW"
        for n, kw in sorted(r.phase_kw.items()))
    print(f"  [{label}]  Vmin={r.min_v:.1f}V  TR={r.tr_loading_pct:.1f}%  "
          f"Imbal={r.phase_imbalance_pct:.1f}%  LowV={len(r.low_v_nodes)}nodes"
          + (f"  [{phase_str}]" if phase_str else ""))


def print_console_report(opt: LVOptimizer) -> None:
    col = 72
    net = opt.net
    b   = opt.baseline
    print(f"\n{'='*col}")
    print(f"  PHASE OPTIMIZER — {net.facilityid}  "
          f"(rated={net.rated_kva:.0f} kVA  meters={len(net.load_feat_nodes)})")
    print(f"{'='*col}")

    print("\nBaseline:")
    _print_sim_line("", b)
    if b.phase_kw:
        for node, kw in sorted(b.phase_kw.items()):
            print(f"  Phase {PHASE_MAP.get(node,'?')}: {kw:.2f} kW")

    if opt.phase_moves:
        pr = opt.phase_result
        print(f"\nPhase Transfer ({len(opt.phase_moves)} moves):")
        for m in opt.phase_moves:
            print(f"  {m.meter.peano or 'idx='+str(m.meter.feat_idx):>14}  "
                  f"{PD_TO_PHASE.get(m.from_pd,'?')}→{PD_TO_PHASE.get(m.to_pd,'?')}  "
                  f"ΔV={m.delta_min_v:+.1f}V  Δimbal={m.delta_imbalance:+.1f}%")
        if pr:
            print(f"  ผลรวม: Vmin={pr.min_v:.1f}V  "
                  f"Imbal={pr.phase_imbalance_pct:.1f}%  LowV={len(pr.low_v_nodes)}")
    else:
        print("\nPhase Transfer: ไม่มีการย้าย")

    if opt.upgrade_options:
        print(f"\nConductor Upgrade (top {min(3, len(opt.upgrade_options))}):")
        for i, up in enumerate(opt.upgrade_options[:3], 1):
            status = (f"Vmin={up.result.min_v:.1f}V  ΔV={up.delta_min_v:+.1f}V"
                      if up.result.converged else "[NG]")
            marker = " <- APPLIED" if (up is opt.applied_upgrade) else ""
            print(f"  #{i} edge ({up.edge[0]}→{up.edge[1]})  "
                  f"{up.from_size}→{up.to_size}mm²  {up.n_affected}LowV  {status}{marker}")
    else:
        print("\nConductor Upgrade: ไม่มีตัวเลือก")

    if opt.applied_upgrade and opt.upgrade_result:
        ur = opt.upgrade_result
        print(f"\nPhase Transfer หลัง CU ({len(opt.phase_moves_cu)} moves):")
        for m in opt.phase_moves_cu:
            print(f"  {m.meter.peano or 'idx='+str(m.meter.feat_idx):>14}  "
                  f"{PD_TO_PHASE.get(m.from_pd,'?')}→{PD_TO_PHASE.get(m.to_pd,'?')}  "
                  f"ΔV={m.delta_min_v:+.1f}V  Δimbal={m.delta_imbalance:+.1f}%")
        if not opt.phase_moves_cu:
            print("  ไม่มีการย้าย")
        if opt.phase_result_cu:
            pur = opt.phase_result_cu
            print(f"  ผลรวม: Vmin={pur.min_v:.1f}V  "
                  f"Imbal={pur.phase_imbalance_pct:.1f}%  LowV={len(pur.low_v_nodes)}")

    if opt.phase_add_options:
        print(f"\nPhase Addition (top {min(3, len(opt.phase_add_options))}):")
        for i, pa in enumerate(opt.phase_add_options[:3], 1):
            status = (f"Vmin={pa.result.min_v:.1f}V  ΔV={pa.delta_min_v:+.1f}V"
                      if pa.result.converged else "[NG]")
            marker = " <- APPLIED" if (pa is opt.applied_phase_add) else ""
            n_up = len(pa.upgraded_edges) if pa.upgraded_edges else 1
            print(f"  #{i} component ({n_up} edges)  "
                  f"->ABC  {len(pa.meter_moves)}meters  {status}{marker}")
    else:
        print("\nPhase Addition: ไม่มีตัวเลือก")

    if opt.phase_moves2:
        pr2 = opt.phase_result2
        print(f"\nPhase Transfer รอบ 2 ({len(opt.phase_moves2)} moves หลัง Phase Addition):")
        for m in opt.phase_moves2:
            print(f"  {m.meter.peano or 'idx='+str(m.meter.feat_idx):>14}  "
                  f"{PD_TO_PHASE.get(m.from_pd,'?')}→{PD_TO_PHASE.get(m.to_pd,'?')}  "
                  f"ΔV={m.delta_min_v:+.1f}V  Δimbal={m.delta_imbalance:+.1f}%")
        if pr2:
            print(f"  ผลรวม: Vmin={pr2.min_v:.1f}V  "
                  f"Imbal={pr2.phase_imbalance_pct:.1f}%  LowV={len(pr2.low_v_nodes)}")

    if (opt.applied_phase_add or opt.applied_upgrade) and opt.final_result:
        fr = opt.final_result
        _steps = []
        if opt.phase_moves:       _steps.append("PT1")
        if opt.applied_upgrade:
            _cu = opt.applied_upgrade
            _steps.append(f"CU({_cu.from_size}→{_cu.to_size}mm²)")
        if opt.phase_moves_cu:    _steps.append("PT_CU")
        if opt.applied_phase_add:
            _pa = opt.applied_phase_add
            _steps.append(f"PA({len(_pa.upgraded_edges)}edges→ABC)")
        if opt.phase_moves2:      _steps.append("PT2")
        steps = " + ".join(_steps) if _steps else "—"
        print(f"\n{'─'*col}")
        print(f"  COMBINED FINAL  ({steps})")
        print(f"{'─'*col}")
        _print_sim_line("", fr)
        if fr.phase_kw:
            for node, kw in sorted(fr.phase_kw.items()):
                print(f"  Phase {PHASE_MAP.get(node,'?')}: {kw:.2f} kW")
        problems = []
        if fr.low_v_nodes:
            problems.append(f"LowV={len(fr.low_v_nodes)} โหนด")
        if fr.phase_imbalance_pct > opt.max_imbalance_pct:
            problems.append(f"Imbal={fr.phase_imbalance_pct:.1f}%>{opt.max_imbalance_pct}%")
        print(f"  สถานะ: {'ผ่านเกณฑ์ ✓' if not problems else 'ยังเกินเกณฑ์: ' + ', '.join(problems)}")

    print(f"\n{'='*col}")


# ─────────────────────────────────────────────────────────────────────────────
# Excel report
# ─────────────────────────────────────────────────────────────────────────────

def save_excel_report(opt: LVOptimizer, out_path: str) -> None:
    wb = openpyxl.Workbook()
    HDR  = PatternFill("solid", fgColor="1F497D")
    HFNT = Font(color="FFFFFF", bold=True)
    OK   = PatternFill("solid", fgColor="E2EFDA")
    NG   = PatternFill("solid", fgColor="FCE4D6")
    WARN = PatternFill("solid", fgColor="FFEB9C")

    def _hdr(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.fill = HDR; c.font = HFNT

    def _autowidth(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(10, w + 2), 36)

    net = opt.net
    b   = opt.baseline
    pr  = opt.phase_result

    # ── Summary ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ["FACILITYID",        net.facilityid],
        ["Transformer kVA",   net.rated_kva],
        ["Network nodes",     len(net.G.nodes())],
        ["Network edges",     len(net.G.edges())],
        ["Total meters",      len(net.load_feat_nodes)],
        [""],
        ["=== Baseline ===", ""],
        ["Vmin (V)",          round(b.min_v, 1)         if b else ""],
        ["TR loading (%)",    round(b.tr_loading_pct, 1) if b else ""],
        ["Phase imbalance (%)".replace("%","%%"), round(b.phase_imbalance_pct, 1) if b else ""],
        ["Low-voltage nodes", len(b.low_v_nodes)         if b else ""],
    ]
    if b and b.phase_kw:
        rows.append([""])
        for node, kw in sorted(b.phase_kw.items()):
            rows.append([f"Phase {PHASE_MAP.get(node,'?')} (kW)", round(kw, 2)])

    if opt.phase_moves and pr:
        rows += [
            [""],
            ["=== After Phase Transfer ===", ""],
            ["Vmin (V)",             round(pr.min_v, 1)],
            ["Phase imbalance (%%)". replace("%%","%"), round(pr.phase_imbalance_pct, 1)],
            ["Low-voltage nodes",    len(pr.low_v_nodes)],
            ["Meters moved",         len(opt.phase_moves)],
        ]

    if opt.upgrade_options:
        best = opt.upgrade_options[0]
        _cu_label = "APPLIED" if (opt.applied_upgrade and opt.applied_upgrade is best) else "ตัวเลือก"
        rows += [
            [""],
            [f"=== Conductor Upgrade ({_cu_label}) ===", ""],
            ["Edge",                f"{best.edge[0]} → {best.edge[1]}"],
            ["Conductor upgrade",   f"{best.from_size} → {best.to_size} mm²"],
            ["Affected LowV nodes", best.n_affected],
            ["ΔVmin sim (V)",       round(best.delta_min_v, 1)],
            ["Sim Vmin (V)",        round(best.result.min_v, 1) if best.result.converged else "NG"],
        ]
        if opt.applied_upgrade and opt.upgrade_result and opt.upgrade_result.converged:
            rows += [
                ["After Apply Vmin (V)",  round(opt.upgrade_result.min_v, 1)],
                ["After Apply Imbal (%)", round(opt.upgrade_result.phase_imbalance_pct, 1)],
                ["PT หลัง CU (moves)",   len(opt.phase_moves_cu)],
            ]
            if opt.phase_result_cu and opt.phase_result_cu.converged:
                rows.append(["After CU+PT Vmin (V)", round(opt.phase_result_cu.min_v, 1)])

    if opt.phase_add_options:
        best = opt.phase_add_options[0]
        rows += [
            [""],
            ["=== Best Phase Addition ===", ""],
            ["Edge",                    f"{best.edge[0]} → {best.edge[1]}"],
            ["Phase upgrade",           f"{_PD_LABEL.get(best.from_pd,'?')} → ABC"],
            ["Meters redistributed",    len(best.meter_moves)],
            ["Affected LowV nodes",     best.n_affected],
            ["ΔVmin (V)",               round(best.delta_min_v, 1)],
            ["After Vmin (V)", round(best.result.min_v, 1) if best.result.converged else "NG"],
        ]

    for row in rows:
        ws.append(row)
    _autowidth(ws)

    # ── Phase Transfer ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Phase Transfer")
    _hdr(ws2, ["#", "PEANO", "NodeID", "From Phase", "To Phase",
               "ΔVmin (V)", "Δimbal (%)"])
    for i, m in enumerate(opt.phase_moves, 1):
        ws2.append([
            i,
            m.meter.peano or f"idx={m.meter.feat_idx}",
            m.meter.node_id,
            PD_TO_PHASE.get(m.from_pd, str(m.from_pd)),
            PD_TO_PHASE.get(m.to_pd,   str(m.to_pd)),
            round(m.delta_min_v, 2),
            round(m.delta_imbalance, 2),
        ])
        fill = OK if m.delta_min_v >= 0 else NG
        for c in ws2[ws2.max_row]: c.fill = fill
    _autowidth(ws2)

    # ── Conductor Upgrade ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Conductor Upgrade")
    _hdr(ws3, ["#", "Edge (u→v)", "From mm²", "To mm²",
               "LowV nodes affected", "ΔVmin (V)", "After Vmin (V)", "Converged"])
    for i, up in enumerate(opt.upgrade_options, 1):
        ws3.append([
            i, f"{up.edge[0]}→{up.edge[1]}",
            up.from_size, up.to_size, up.n_affected,
            round(up.delta_min_v, 1),
            round(up.result.min_v, 1) if up.result.converged else "",
            "YES" if up.result.converged else "NO",
        ])
        fill = OK if up.delta_min_v > 0 else NG
        for c in ws3[ws3.max_row]: c.fill = fill
    _autowidth(ws3)

    # ── Phase Addition ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Phase Addition")
    _hdr(ws4, ["#", "Edge (u→v)", "From phases", "To phases",
               "Meters moved", "LowV nodes affected",
               "ΔVmin (V)", "After Vmin (V)", "Converged"])
    for i, pa in enumerate(opt.phase_add_options, 1):
        ws4.append([
            i, f"{pa.edge[0]}→{pa.edge[1]}",
            _PD_LABEL.get(pa.from_pd, str(pa.from_pd)), "ABC",
            len(pa.meter_moves), pa.n_affected,
            round(pa.delta_min_v, 1),
            round(pa.result.min_v, 1) if pa.result.converged else "",
            "YES" if pa.result.converged else "NO",
        ])
        fill = OK if pa.delta_min_v > 0 else NG
        for c in ws4[ws4.max_row]: c.fill = fill
    _autowidth(ws4)

    # ── Meter Detail (all meters) ─────────────────────────────────────────
    ws5 = wb.create_sheet("Meters")
    _hdr(ws5, ["FeatIdx", "PEANO", "NodeID", "Phase final (PD)", "Conductor PD",
               "Conductor phases", "Load kW", "Movable?"])
    _raw_for_meters = opt.final_raw if opt.final_raw is not None else opt.raw
    for m in _build_meter_inventory(opt.net, _raw_for_meters):
        ws5.append([
            m.feat_idx, m.peano, m.node_id,
            PD_TO_PHASE.get(m.current_pd, str(m.current_pd)),
            m.conductor_pd, _PD_LABEL.get(m.conductor_pd, str(m.conductor_pd)),
            round(m.kw, 3),
            "YES" if (m.is_single_phase and m.movable_pds) else "NO",
        ])
    _autowidth(ws5)

    wb.save(out_path)
    print(f"[Excel] บันทึก: {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Static PNG map
# ─────────────────────────────────────────────────────────────────────────────

def draw_map(opt: LVOptimizer, out_path: str) -> None:
    net  = opt.net
    b    = opt.baseline
    # ใช้ผลลัพธ์สุดท้าย (final_result) แสดงแรงดัน — ถ้าไม่มีใช้ baseline
    display_r = opt.final_result if opt.final_result else b
    node_v = {nid: min(ph.values()) for nid, ph in display_r.node_phase_v.items() if ph}
    has_v  = bool(node_v)

    norm  = mcolors.Normalize(vmin=150, vmax=240)
    cmap  = mcolors.LinearSegmentedColormap.from_list(
        "volt", ["#d73027", "#fee08b", "#91cf60", "#1a9850"], N=256)
    BASE  = "#4472C4"

    fig, ax = plt.subplots(figsize=(15, 10))

    # Edges
    for u, v in net.G.edges():
        xu, yu = net.node_coords[u]; xv, yv = net.node_coords[v]
        ax.plot([xu, xv], [yu, yv], color=BASE, lw=1.2, alpha=0.45, zorder=2)

    # Nodes coloured by final voltage
    for nid, (x, y) in net.node_coords.items():
        s = 22 + net.node_kw.get(nid, 0) * 1.4
        v = node_v.get(nid)
        c = cmap(norm(v)) if v is not None else BASE
        ax.scatter(x, y, color=c, s=s, zorder=3, alpha=0.85)

    # Red ring — โหนดที่ยังมีแรงดันต่ำหลัง optimize (final state)
    for nid in display_r.low_v_nodes:
        if nid in net.node_coords:
            x, y = net.node_coords[nid]
            ax.scatter(x, y, s=75, facecolors="none", edgecolors="red",
                       linewidths=1.5, zorder=4)

    # Green ring for meters that were moved (all PT rounds)
    moved_nids = (
        {m.meter.node_id for m in opt.phase_moves} |
        {m.meter.node_id for m in opt.phase_moves_cu} |
        {m.meter.node_id for m in opt.phase_moves2}
    )
    for nid in moved_nids:
        if nid in net.node_coords:
            x, y = net.node_coords[nid]
            ax.scatter(x, y, s=110, facecolors="none", edgecolors="#00B050",
                       linewidths=2.0, zorder=5)

    # Orange dashed — best conductor upgrade edge
    legend_extra: List = []
    if opt.upgrade_options:
        up = opt.upgrade_options[0]
        u2, v2 = up.edge
        if u2 in net.node_coords and v2 in net.node_coords:
            xu2, yu2 = net.node_coords[u2]; xv2, yv2 = net.node_coords[v2]
            ax.plot([xu2, xv2], [yu2, yv2], color="#FF6600", lw=4,
                    linestyle="--", zorder=6)
            legend_extra.append(
                Line2D([0],[0], color="#FF6600", lw=3, linestyle="--",
                       label=f"แนะนำ: เพิ่มขนาดสาย {up.from_size}→{up.to_size}mm²"))

    # Purple dash-dot — best phase addition: draw ALL upgraded edges (ไม่ใช่แค่ edge เดียว)
    if opt.phase_add_options:
        pa = opt.phase_add_options[0]
        pa_xs, pa_ys = [], []
        for _fi, eu, ev in pa.upgraded_edges:
            if eu in net.node_coords and ev in net.node_coords:
                xu2, yu2 = net.node_coords[eu]; xv2, yv2 = net.node_coords[ev]
                pa_xs += [xu2, xv2, None]; pa_ys += [yu2, yv2, None]
        if pa_xs:
            ax.plot(pa_xs, pa_ys, color="#8B00FF", lw=3,
                    linestyle="-.", zorder=6, alpha=0.85)
            legend_extra.append(
                Line2D([0],[0], color="#8B00FF", lw=3, linestyle="-.",
                       label=f"แนะนำ: เพิ่มเฟส→ABC ({len(pa.upgraded_edges)} edges)"))

    # Transformer star
    tx, ty = net.node_coords[net.transformer_node]
    ax.scatter(tx, ty, color=BASE, marker="*", s=480, zorder=7,
               edgecolors="k", linewidths=0.8)
    ax.annotate(f"TR\n{net.facilityid}", (tx, ty),
                xytext=(7, 7), textcoords="offset points",
                fontsize=8, color=BASE, fontweight="bold")

    # Colorbar
    if has_v:
        sm = mcm.ScalarMappable(cmap=cmap, norm=norm)
        sm.set_array([])
        cbar = fig.colorbar(sm, ax=ax, shrink=0.58, pad=0.02)
        cbar.set_label("Voltage (V)", fontsize=9)
        y_thr = (opt.min_voltage_v - 150) / 90
        cbar.ax.axhline(y=y_thr, color="red", lw=1.5, linestyle="--")
        cbar.ax.text(1.1, y_thr, f"{opt.min_voltage_v:.0f}V",
                     transform=cbar.ax.transAxes, fontsize=7, color="red", va="center")

    legend_items = [
        Line2D([0],[0], color=BASE, lw=2,        label=net.facilityid),
        Line2D([0],[0], marker="*", color=BASE, markersize=12, lw=0, label="หม้อแปลง"),
        Line2D([0],[0], marker="o", color="none", markerfacecolor="none",
               markeredgecolor="red", markersize=8, markeredgewidth=1.5, lw=0,
               label=f"V < {opt.min_voltage_v:.0f}V (หลัง optimize)"),
        Line2D([0],[0], marker="o", color="none", markerfacecolor="none",
               markeredgecolor="#00B050", markersize=8, markeredgewidth=2, lw=0,
               label="มิเตอร์ที่ย้ายเฟส"),
    ] + legend_extra

    ax.legend(handles=legend_items, loc="upper right", fontsize=8, framealpha=0.9)
    # Title: baseline + all stages
    title = (f"LV Phase/Conductor Analysis  |  {net.facilityid}  "
             f"(rated={net.rated_kva:.0f}kVA)\n"
             f"Baseline: Vmin={b.min_v:.1f}V  TR={b.tr_loading_pct:.1f}%  "
             f"Imbal={b.phase_imbalance_pct:.1f}%  LowV={len(b.low_v_nodes)}nodes")
    if opt.final_result and opt.final_result is not b:
        fr = opt.final_result
        title += (f"\nFinal: Vmin={fr.min_v:.1f}V  TR={fr.tr_loading_pct:.1f}%  "
                  f"Imbal={fr.phase_imbalance_pct:.1f}%  LowV={len(fr.low_v_nodes)}nodes"
                  f"  (สี node = แรงดันหลัง optimize)")

    ax.set_title(title, fontsize=10)
    ax.set_xlabel("Easting (UTM47N, m)")
    ax.set_ylabel("Northing (UTM47N, m)")
    ax.ticklabel_format(style="plain", axis="both")
    ax.grid(True, alpha=0.22)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"[Map]   PNG: {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Interactive Plotly map
# ─────────────────────────────────────────────────────────────────────────────

def draw_interactive_map(opt: LVOptimizer, out_path: str) -> None:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    net   = opt.net
    b     = opt.baseline
    pr    = opt.phase_result
    # แสดงแรงดันจาก final_result (หลัง optimize ทุกขั้นตอน) ถ้ามี
    display_r = opt.final_result if opt.final_result else b
    node_v = {nid: min(ph.values()) for nid, ph in display_r.node_phase_v.items() if ph}
    vmin_c, vmax_c = 150, 240
    cscale = [[0.0, "#d73027"],[0.33, "#fee08b"],[0.66, "#91cf60"],[1.0, "#1a9850"]]

    moved_nids = (
        {m.meter.node_id for m in opt.phase_moves} |
        {m.meter.node_id for m in opt.phase_moves_cu} |
        {m.meter.node_id for m in opt.phase_moves2}
    )
    upgrade_edges = {up.edge for up in opt.upgrade_options[:1]}

    # Phase colour for edge (by majority conductor PD)
    _phase_color = {7:"#888888", 6:"#0070C0", 5:"#7030A0", 4:"#FF0000",
                    3:"#00B050", 2:"#0070C0", 1:"#7030A0"}

    fig = make_subplots(
        rows=2, cols=1,
        row_heights=[0.70, 0.30],
        specs=[[{"type":"scatter"}],[{"type":"table"}]],
        vertical_spacing=0.04,
    )

    # ── Edges (colored by phase) ─────────────────────────────────────────
    # ใช้ BFS tree edges เท่านั้น — lc_feat_edges มี loop edges ด้วย ทำให้เห็น 2 เส้น
    # สร้าง reverse lookup: (u,v) หรือ (v,u) → feat_idx สำหรับ PHASEDESIGNATION
    _edge_feat: Dict[Tuple[int,int], int] = {}
    for fi, (fu, fv) in net.lc_feat_edges.items():
        _edge_feat[(fu, fv)] = fi
        _edge_feat[(fv, fu)] = fi

    pd_groups: Dict[int, Tuple[List, List]] = defaultdict(lambda: ([],[]))
    for u, v in net.G.edges():          # BFS tree edges only (no duplicates)
        if u not in net.node_coords or v not in net.node_coords:
            continue
        feat_idx = _edge_feat.get((u, v)) or _edge_feat.get((v, u))
        if feat_idx is not None:
            pd = int(get_attr(opt.raw["features"][feat_idx], "PHASEDESIGNATION", 7) or 7)
        else:
            pd = 7
        xs, ys = pd_groups[pd]
        xu, yu = net.node_coords[u]; xv, yv = net.node_coords[v]
        xs += [xu, xv, None]; ys += [yu, yv, None]

    for pd_val, (xs, ys) in pd_groups.items():
        color = _phase_color.get(pd_val, "#888888")
        fig.add_trace(go.Scatter(
            x=xs, y=ys, mode="lines",
            line=dict(color=color, width=1.8), opacity=0.6,
            name=f"สาย {_PD_LABEL.get(pd_val,'?')}",
            hoverinfo="skip",
        ), row=1, col=1)

    # ── Nodes coloured by voltage ─────────────────────────────────────────
    nx_v, ny_v, nv_v, nt_v, ns_v = [], [], [], [], []
    nx_n, ny_n, nt_n, ns_n = [], [], [], []

    def _phase_tip(phases):
        return "<br>" + "<br>".join(
            f"V{PHASE_MAP.get(n,'?')} = {v:.1f} V"
            for n, v in sorted(phases.items())) if phases else ""

    for nid, (x, y) in net.node_coords.items():
        kw   = net.node_kw.get(nid, 0.0)
        v    = node_v.get(nid)
        ph   = b.node_phase_v.get(nid, {})
        sz   = 7 + min(kw * 0.6, 14)
        tip  = f"<b>nid={nid}</b>" + _phase_tip(ph) + f"<br>Load={kw:.2f}kW"
        if v is not None:
            nx_v.append(x); ny_v.append(y); nv_v.append(v)
            nt_v.append(tip + f"<br><b>Vmin={v:.1f}V</b>"); ns_v.append(sz)
        else:
            nx_n.append(x); ny_n.append(y)
            nt_n.append(tip); ns_n.append(sz)

    if nx_v:
        fig.add_trace(go.Scatter(
            x=nx_v, y=ny_v, mode="markers",
            marker=dict(color=nv_v, colorscale=cscale, cmin=vmin_c, cmax=vmax_c,
                        size=ns_v, opacity=0.88, line=dict(width=0),
                        colorbar=dict(title="V (V)", thickness=14, len=0.50,
                                      x=1.01, outlinewidth=0.5,
                                      tickvals=list(range(150, 241, 10)))),
            name="โหนด (แรงดัน)", text=nt_v,
            hovertemplate="%{text}<extra></extra>",
        ), row=1, col=1)
    if nx_n:
        fig.add_trace(go.Scatter(
            x=nx_n, y=ny_n, mode="markers",
            marker=dict(color="#4472C4", size=ns_n, opacity=0.5),
            name="โหนด (ไม่มีข้อมูล)", text=nt_n,
            hovertemplate="%{text}<extra></extra>",
        ), row=1, col=1)

    # ── Low-voltage rings (แสดงเฉพาะโหนดที่ยังต่ำหลัง optimize) ───────────
    lx, ly, lt = [], [], []
    for nid in display_r.low_v_nodes:
        if nid in net.node_coords:
            x, y = net.node_coords[nid]
            v = node_v.get(nid, 0)
            ph = display_r.node_phase_v.get(nid, {})
            lx.append(x); ly.append(y)
            lt.append(f"<b>LowV nid={nid}  Vmin={v:.1f}V</b>" + _phase_tip(ph))
    if lx:
        fig.add_trace(go.Scatter(
            x=lx, y=ly, mode="markers",
            marker=dict(symbol="circle-open", color="red", size=18, line=dict(width=2.5)),
            name=f"V < {opt.min_voltage_v:.0f}V",
            text=lt, hovertemplate="%{text}<extra></extra>",
        ), row=1, col=1)

    # ── Moved meters (green diamond) ─────────────────────────────────────
    mx, my, mt = [], [], []
    for m in opt.phase_moves:
        if m.meter.node_id in net.node_coords:
            x, y = net.node_coords[m.meter.node_id]
            mx.append(x); my.append(y)
            mt.append(f"<b>ย้ายเฟส: {m.meter.peano or 'idx='+str(m.meter.feat_idx)}</b>"
                      f"<br>{PD_TO_PHASE.get(m.from_pd,'?')} → "
                      f"{PD_TO_PHASE.get(m.to_pd,'?')}"
                      f"<br>ΔV={m.delta_min_v:+.1f}V  kW={m.meter.kw:.2f}")
    if mx:
        fig.add_trace(go.Scatter(
            x=mx, y=my, mode="markers",
            marker=dict(symbol="circle-open", color="#00B050", size=20,
                        line=dict(width=2.5)),
            name="มิเตอร์ที่ย้ายเฟส",
            text=mt, hovertemplate="%{text}<extra></extra>",
        ), row=1, col=1)

    # ── Conductor upgrade edge ────────────────────────────────────────────
    for up in opt.upgrade_options[:1]:
        u2, v2 = up.edge
        if u2 in net.node_coords and v2 in net.node_coords:
            xu2, yu2 = net.node_coords[u2]; xv2, yv2 = net.node_coords[v2]
            tip = (f"<b>แนะนำ: เพิ่มขนาดสาย</b><br>"
                   f"({u2}→{v2})  {up.from_size}→{up.to_size}mm²<br>"
                   f"ΔV={up.delta_min_v:+.1f}V<extra></extra>")
            fig.add_trace(go.Scatter(
                x=[xu2, xv2], y=[yu2, yv2], mode="lines+markers",
                line=dict(color="#FF6600", width=5, dash="dash"),
                marker=dict(symbol="diamond", size=12, color="#FF6600"),
                name=f"เพิ่มขนาดสาย {up.from_size}→{up.to_size}mm²",
                hovertemplate=tip,
            ), row=1, col=1)

    # ── Phase addition: draw ALL upgraded edges (ไม่ใช่แค่ edge เดียว) ─────
    for pa in opt.phase_add_options[:1]:
        pa_xs, pa_ys = [], []
        for _fi, eu, ev in pa.upgraded_edges:
            if eu in net.node_coords and ev in net.node_coords:
                xu2, yu2 = net.node_coords[eu]; xv2, yv2 = net.node_coords[ev]
                pa_xs += [xu2, xv2, None]; pa_ys += [yu2, yv2, None]
        if pa_xs:
            tip = (f"<b>แนะนำ: เพิ่มเฟสสาย→ABC</b><br>"
                   f"{len(pa.upgraded_edges)} edges<br>"
                   f"ย้าย {len(pa.meter_moves)} มิเตอร์<br>"
                   f"ΔV={pa.delta_min_v:+.1f}V<extra></extra>")
            fig.add_trace(go.Scatter(
                x=pa_xs, y=pa_ys, mode="lines",
                line=dict(color="#8B00FF", width=4, dash="dashdot"),
                name=f"เพิ่มเฟส→ABC ({len(pa.upgraded_edges)} edges)",
                hovertemplate=tip,
                opacity=0.85,
            ), row=1, col=1)

    # ── Meter points ──────────────────────────────────────────────────────
    mmx, mmy, mmv, mmt = [], [], [], []
    for feat_idx, node_id in net.load_feat_nodes.items():
        feat = opt.raw["features"][feat_idx]
        g = feat.get("geometry", {})
        fx, fy = g.get("x"), g.get("y")
        if fx is None or fy is None:
            continue
        peano = str(get_attr(feat, "PEANO", "") or "").strip()
        kw    = float(get_attr(feat, "KWP", 0.0) or 0.0)
        v     = node_v.get(node_id)
        tip   = f"<b>Meter {peano or 'nid='+str(node_id)}</b><br>kW={kw:.2f}<br>"
        tip  += f"<b>V={v:.1f}V</b>" if v else "V=—"
        mmx.append(fx); mmy.append(fy)
        mmv.append(v if v else float("nan"))
        mmt.append(tip)

    vx2, vy2, vv2, vt2 = [], [], [], []
    nx2, ny2, nt2       = [], [], []
    for x, y, v, t in zip(mmx, mmy, mmv, mmt):
        if not math.isnan(v):
            vx2.append(x); vy2.append(y); vv2.append(v); vt2.append(t)
        else:
            nx2.append(x); ny2.append(y); nt2.append(t)
    if vx2:
        fig.add_trace(go.Scatter(
            x=vx2, y=vy2, mode="markers",
            marker=dict(symbol="diamond", color=vv2, colorscale=cscale,
                        cmin=vmin_c, cmax=vmax_c, size=9, opacity=0.88,
                        line=dict(width=0.8, color="rgba(0,0,0,0.3)"), showscale=False),
            name="มิเตอร์ลูกค้า", text=vt2,
            hovertemplate="%{text}<extra></extra>",
        ), row=1, col=1)
    if nx2:
        fig.add_trace(go.Scatter(
            x=nx2, y=ny2, mode="markers",
            marker=dict(symbol="diamond", color="gray", size=7, opacity=0.45),
            name="มิเตอร์ (ไม่มี V)", text=nt2,
            hovertemplate="%{text}<extra></extra>",
        ), row=1, col=1)

    # ── Transformer ───────────────────────────────────────────────────────
    tx, ty = net.node_coords[net.transformer_node]
    fig.add_trace(go.Scatter(
        x=[tx], y=[ty], mode="markers+text",
        marker=dict(symbol="star", size=24, color="#4472C4",
                    line=dict(color="black", width=1.5)),
        text=[f"TR {net.facilityid}"], textposition="top right",
        textfont=dict(size=10, color="#4472C4"),
        name=f"หม้อแปลง {net.facilityid}",
        hovertemplate=(f"<b>TR {net.facilityid}</b><br>"
                       f"rated={net.rated_kva:.0f}kVA<br>"
                       f"({tx:.0f},{ty:.0f})<extra></extra>"),
    ), row=1, col=1)

    # ── Summary table ─────────────────────────────────────────────────────
    rows_hdr = ["ขั้นตอน", "รายละเอียด", "Vmin (V)", "Imbal (%)", "LowV nodes", "ΔV (V)"]
    rows_val: List[List] = []
    rows_clr: List[str]  = []

    rows_val.append(["Baseline", f"TR={b.tr_loading_pct:.1f}%",
                     f"{b.min_v:.1f}", f"{b.phase_imbalance_pct:.1f}",
                     str(len(b.low_v_nodes)), "—"])
    rows_clr.append("#fff2cc")

    if opt.phase_moves and pr:
        rows_val.append([f"Phase Transfer ({len(opt.phase_moves)} moves)",
                         ", ".join(f"{PD_TO_PHASE.get(m.from_pd,'?')}→{PD_TO_PHASE.get(m.to_pd,'?')}"
                                   for m in opt.phase_moves[:4]) + ("…" if len(opt.phase_moves)>4 else ""),
                         f"{pr.min_v:.1f}", f"{pr.phase_imbalance_pct:.1f}",
                         str(len(pr.low_v_nodes)),
                         f"{pr.min_v - b.min_v:+.1f}"])
        rows_clr.append("#e2efda" if pr.min_v > b.min_v else "#fce4d6")

    if opt.applied_upgrade and opt.upgrade_result and opt.upgrade_result.converged:
        _cu = opt.applied_upgrade
        ur = opt.upgrade_result
        rows_val.append([f"CU Apply {_cu.from_size}→{_cu.to_size}mm² ✓",
                         f"edge ({_cu.edge[0]}→{_cu.edge[1]})  {_cu.n_affected}LowV",
                         f"{ur.min_v:.1f}", f"{ur.phase_imbalance_pct:.1f}",
                         str(len(ur.low_v_nodes)), f"{_cu.delta_min_v:+.1f}"])
        rows_clr.append("#e2efda" if _cu.delta_min_v > 0 else "#fce4d6")

    if opt.phase_moves_cu and opt.phase_result_cu and opt.phase_result_cu.converged:
        pur = opt.phase_result_cu
        ur_ref = opt.upgrade_result
        dv_cu = (pur.min_v - ur_ref.min_v) if ur_ref and ur_ref.converged else 0.0
        rows_val.append([f"PT หลัง CU ({len(opt.phase_moves_cu)} moves)",
                         ", ".join(f"{PD_TO_PHASE.get(m.from_pd,'?')}→{PD_TO_PHASE.get(m.to_pd,'?')}"
                                   for m in opt.phase_moves_cu[:4])
                         + ("…" if len(opt.phase_moves_cu) > 4 else ""),
                         f"{pur.min_v:.1f}", f"{pur.phase_imbalance_pct:.1f}",
                         str(len(pur.low_v_nodes)), f"{dv_cu:+.1f}"])
        rows_clr.append("#e2efda" if dv_cu > 0 else "#fce4d6")

    for up in opt.upgrade_options[:2]:
        r2 = up.result
        marker = " ✓" if (up is opt.applied_upgrade) else ""
        rows_val.append([f"Conductor {up.from_size}→{up.to_size}mm²{marker}",
                         f"edge ({up.edge[0]}→{up.edge[1]})  {up.n_affected}โหนด",
                         f"{r2.min_v:.1f}" if r2.converged else "NG",
                         f"{r2.phase_imbalance_pct:.1f}" if r2.converged else "—",
                         str(len(r2.low_v_nodes)) if r2.converged else "—",
                         f"{up.delta_min_v:+.1f}"])
        rows_clr.append("#e2efda" if up.delta_min_v > 0 else "#fce4d6")

    for pa in opt.phase_add_options[:2]:
        r2 = pa.result
        rows_val.append([f"Phase Addition {_PD_LABEL.get(pa.from_pd,'?')}→ABC",
                         f"edge ({pa.edge[0]}→{pa.edge[1]})  {len(pa.meter_moves)}มิเตอร์",
                         f"{r2.min_v:.1f}" if r2.converged else "NG",
                         f"{r2.phase_imbalance_pct:.1f}" if r2.converged else "—",
                         str(len(r2.low_v_nodes)) if r2.converged else "—",
                         f"{pa.delta_min_v:+.1f}"])
        rows_clr.append("#e2efda" if pa.delta_min_v > 0 else "#fce4d6")

    cols_t = list(zip(*rows_val)) if rows_val else [[] for _ in rows_hdr]
    fig.add_trace(go.Table(
        header=dict(values=rows_hdr,
                    fill_color="#1F497D", font=dict(color="white", size=11),
                    align="center", height=28),
        cells=dict(values=cols_t,
                   fill_color=[rows_clr] * len(rows_hdr),
                   font=dict(size=11), align="center", height=24),
    ), row=2, col=1)

    # ── Layout ────────────────────────────────────────────────────────────
    fig.update_layout(
        title=dict(
            text=(f"LV Phase/Conductor Analysis  |  {net.facilityid}  "
                  f"(rated={net.rated_kva:.0f}kVA)<br>"
                  f"<sup>Baseline: Vmin={b.min_v:.1f}V  Imbal={b.phase_imbalance_pct:.1f}%  "
                  f"LowV={len(b.low_v_nodes)}"
                  + (f"  →  Final: Vmin={display_r.min_v:.1f}V  "
                     f"Imbal={display_r.phase_imbalance_pct:.1f}%  "
                     f"LowV={len(display_r.low_v_nodes)}"
                     if display_r is not b else "")
                  + f"  |  สีโหนด = แรงดันหลัง optimize  |  Click legend = toggle</sup>"),
            font=dict(size=13),
        ),
        hovermode="closest",
        plot_bgcolor="white", paper_bgcolor="white",
        height=1080,
        legend=dict(x=1.07, y=0.99, bgcolor="rgba(255,255,255,0.92)",
                    bordercolor="#cccccc", borderwidth=1, font=dict(size=11)),
        margin=dict(r=220, t=95, l=70, b=30),
    )
    fig.update_xaxes(title_text="Easting (UTM47N, m)", tickformat="d",
                     showgrid=True, gridcolor="#eeeeee", zeroline=False, row=1, col=1)
    fig.update_yaxes(title_text="Northing (UTM47N, m)", tickformat="d",
                     showgrid=True, gridcolor="#eeeeee", zeroline=False,
                     scaleanchor="x", scaleratio=1, row=1, col=1)

    fig.write_html(out_path, include_plotlyjs="cdn")
    print(f"[Map]   Interactive HTML: {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(
        description="LV Phase & Conductor Optimizer — วิเคราะห์และปรับปรุงระบบ LV",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    ap.add_argument("facilityid", nargs="?", default=None,
                    help="FACILITYID หม้อแปลง เช่น 04-123456 (ถ้าไม่ใส่จะถามอีกครั้ง)")
    ap.add_argument("--min-voltage",   type=float, default=200.0, metavar="V",
                    help="แรงดันต่ำสุดที่ยอมรับ (V)")
    ap.add_argument("--max-imbalance", type=float, default=25.0,  metavar="PCT",
                    help="phase imbalance สูงสุด (%%)")
    ap.add_argument("--json-dir",      default=JSON_DIR,
                    help="folder ที่เก็บไฟล์ JSON")
    ap.add_argument("--out-dir",       default=DEFAULT_OUT,
                    help="folder สำหรับบันทึก Excel/PNG/HTML")
    ap.add_argument("--max-sim-time",  type=float, default=300.0, metavar="SEC",
                    help="เวลาสูงสุด (วินาที) สำหรับ Phase Transfer แต่ละรอบ")
    ap.add_argument("--region",        default=None,
                    help="เขต GIS ของ กฟภ. ที่หม้อแปลงอยู่ (เช่น NE1, C2, Z)")
    args = ap.parse_args()

    if not args.facilityid:
        args.facilityid = input("กรุณากรอก FACILITYID (เช่น 04-123456): ").strip()
    if not args.facilityid:
        print("[!] ไม่ได้ระบุ FACILITYID — ยกเลิก")
        sys.exit(1)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    fid  = re.sub(r"[^A-Za-z0-9]", "_", args.facilityid)
    stem = out_dir / f"phase_opt_{fid}_{ts}"

    optimizer = LVOptimizer(
        facilityid=args.facilityid,
        min_voltage_v=args.min_voltage,
        max_imbalance_pct=args.max_imbalance,
        json_dir=args.json_dir,
        max_sim_time_s=args.max_sim_time,
        region=args.region,
    )
    optimizer.run()

    if optimizer.net is None or optimizer.baseline is None:
        print("\n[!] วิเคราะห์ไม่สำเร็จ — ตรวจสอบ JSON / GIS connection")
        sys.exit(1)

    print_console_report(optimizer)
    save_excel_report(optimizer, str(stem) + ".xlsx")
    draw_map(optimizer, str(stem) + ".png")
    draw_interactive_map(optimizer, str(stem) + ".html")

    print(f"\nเสร็จสิ้น — ผลลัพธ์บันทึกที่  {out_dir}")


if __name__ == "__main__":
    main()
