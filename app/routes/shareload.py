import os
import re
import sys
import math
import json
import shutil
import subprocess
import threading
import logging
import openpyxl
from pathlib import Path
from datetime import datetime
from flask import (
    Blueprint, request, render_template, jsonify,
    session, abort, send_file
)

from ..config import Config
from ..utils.decorators import login_required
from ..utils.helpers import get_user_region

shareload_bp = Blueprint('shareload', __name__)

_ROOT = Path(os.path.dirname(__file__)).parent.parent
SHARELOAD_DIR = _ROOT / 'feature_shareload'
FEASIBLE_DIR = SHARELOAD_DIR / 'FEASIBLE'
FAC_RE = re.compile(r'^\d{2}-\d{6}$')
PAIR_RE = re.compile(r'^(\d{2}-\d{6})_(\d{2}-\d{6})$')

_running_jobs = set()
_running_lock = threading.Lock()


# ------------------------------------------------------------------
# Coordinate conversion: WGS84/UTM Zone 47N (EPSG:32647) → lat/lon
# ------------------------------------------------------------------

def _utm32647_to_latlon(x: float, y: float) -> tuple:
    """Convert EPSG:32647 (UTM Zone 47N) easting/northing to (lat, lon) degrees."""
    try:
        from pyproj import Transformer
        t = Transformer.from_crs("EPSG:32647", "EPSG:4326", always_xy=True)
        lon, lat = t.transform(x, y)
        return float(lat), float(lon)
    except Exception:
        pass

    # Manual Bowring/Redfearn UTM → geographic (WGS84)
    a = 6378137.0
    f = 1.0 / 298.257223563
    b = a * (1 - f)
    e2 = (a**2 - b**2) / a**2
    ep2 = (a**2 - b**2) / b**2
    k0 = 0.9996
    E0 = 500000.0
    lon0 = math.radians(99.0)   # Zone 47N central meridian

    xr = x - E0
    yr = y          # northern hemisphere: N0 = 0

    M = yr / k0
    e_ = math.sqrt(1 - e2)
    mu = M / (a * (1 - e2/4 - 3*e2**2/64 - 5*e2**3/256))
    e1 = (1 - e_) / (1 + e_)

    phi1 = (mu
            + (3*e1/2 - 27*e1**3/32) * math.sin(2*mu)
            + (21*e1**2/16 - 55*e1**4/32) * math.sin(4*mu)
            + (151*e1**3/96) * math.sin(6*mu)
            + (1097*e1**4/512) * math.sin(8*mu))

    sp = math.sin(phi1)
    N1 = a / math.sqrt(1 - e2 * sp**2)
    T1 = math.tan(phi1)**2
    C1 = ep2 * math.cos(phi1)**2
    R1 = a * (1 - e2) / (1 - e2 * sp**2)**1.5
    D = xr / (N1 * k0)

    lat = phi1 - (N1 * math.tan(phi1) / R1) * (
        D**2/2
        - (5 + 3*T1 + 10*C1 - 4*C1**2 - 9*ep2) * D**4/24
        + (61 + 90*T1 + 298*C1 + 45*T1**2 - 252*ep2 - 3*C1**2) * D**6/720
    )
    lon = lon0 + (
        D
        - (1 + 2*T1 + C1) * D**3/6
        + (5 - 2*C1 + 28*T1 - 3*C1**2 + 8*ep2 + 24*T1**2) * D**5/120
    ) / math.cos(phi1)

    return math.degrees(lat), math.degrees(lon)


def _ll(x, y):
    """Return [lon, lat] list for ArcGIS (handles None gracefully)."""
    if x is None or y is None:
        return None
    lat, lon = _utm32647_to_latlon(x, y)
    return [round(lon, 7), round(lat, 7)]


# ------------------------------------------------------------------
# Plotly HTML parser
# ------------------------------------------------------------------
#
# The map data is pulled straight out of the Plotly figure's own trace
# array (embedded in the HTML as `Plotly.newPlot("<div>", [<traces>], ...)`)
# rather than regex-chaining "x"/"y"/"color"/"text" keys across the raw
# serialized text. Plotly serializes each trace's keys alphabetically, so
# `marker.color` sorts *before* `name` while `x`/`y`/`text` sort *after*
# it — a forward-only regex chain starting at "name" can walk past the
# end of the current trace and silently pick up a *later* trace's color/
# text arrays once the current trace has no further "color" key of its
# own after "y". Parsing the whole trace array as JSON sidesteps that
# entirely and also gets automatic \uXXXX unescaping for free.

_PLOTLY_DATA_RE = re.compile(r'Plotly\.newPlot\(\s*"[^"]*",\s*(\[)')


def _match_bracket(html: str, start: int) -> int | None:
    """Return the index just past the ']' matching the '[' at `start`."""
    depth = 0
    in_str = False
    esc = False
    for i in range(start, len(html)):
        c = html[i]
        if in_str:
            if esc:
                esc = False
            elif c == '\\':
                esc = True
            elif c == '"':
                in_str = False
            continue
        if c == '"':
            in_str = True
        elif c == '[':
            depth += 1
        elif c == ']':
            depth -= 1
            if depth == 0:
                return i + 1
    return None


def _extract_plotly_traces(html: str) -> list:
    m = _PLOTLY_DATA_RE.search(html)
    if not m:
        return []
    start = m.start(1)
    end = _match_bracket(html, start)
    if end is None:
        return []
    try:
        traces = json.loads(html[start:end])
    except (ValueError, TypeError):
        return []
    return traces if isinstance(traces, list) else []


def _trace_by_name(traces: list, name: str) -> dict | None:
    return next((t for t in traces if t.get('name') == name), None)


def _trace_by_prefix(traces: list, prefix: str) -> dict | None:
    return next((t for t in traces if str(t.get('name') or '').startswith(prefix)), None)


def _lines_to_paths(xs: list, ys: list) -> list:
    """Split null-separated x/y arrays into list of [[lon,lat],...] paths."""
    paths, current = [], []
    for x, y in zip(xs, ys):
        if x is None or y is None:
            if len(current) >= 2:
                paths.append(current)
            current = []
        else:
            pt = _ll(x, y)
            if pt:
                current.append(pt)
    if len(current) >= 2:
        paths.append(current)
    return paths


def _trace_pts(trace: dict) -> list:
    """[lon,lat] points for a trace's x/y arrays (nulls/invalid dropped)."""
    xs, ys = trace.get('x') or [], trace.get('y') or []
    pts = [_ll(x, y) for x, y in zip(xs, ys) if x is not None and y is not None]
    return [p for p in pts if p]


def _voltage_nodes_from_trace(trace: dict) -> list:
    xs, ys = trace.get('x') or [], trace.get('y') or []
    vs = (trace.get('marker') or {}).get('color') or []
    texts = trace.get('text') or []
    nodes = []
    for i, (x, y) in enumerate(zip(xs, ys)):
        if x is None or y is None:
            continue
        pt = _ll(x, y)
        if not pt:
            continue
        label = texts[i] if i < len(texts) else ''
        nodes.append({
            'lon': pt[0], 'lat': pt[1],
            'v': vs[i] if i < len(vs) and isinstance(vs[i], (int, float)) else None,
            'label': str(label).replace('<br>', '\n').replace('<b>', '').replace('</b>', ''),
        })
    return nodes


def _meter_points(traces: list, prefix: str) -> list:
    """Merge points across all meter traces starting with `prefix`.

    The map draws meters as up to two separate traces per network — e.g.
    "TR_A Meters" (colored by simulated voltage) and "TR_A Meters (no sim)"
    (gray, when no optimal/feasible scenario exists to simulate voltages) —
    so an exact-name match would silently drop meters in the no-sim case.
    """
    pts = []
    for t in traces:
        if not str(t.get('name') or '').startswith(prefix):
            continue
        colors = (t.get('marker') or {}).get('color')
        colors = colors if isinstance(colors, list) else None
        for i, (x, y) in enumerate(zip(t.get('x') or [], t.get('y') or [])):
            if x is None or y is None:
                continue
            p = _ll(x, y)
            if not p:
                continue
            v = colors[i] if colors and i < len(colors) and isinstance(colors[i], (int, float)) else None
            pts.append({'lon': p[0], 'lat': p[1], 'v': v})
    return pts


# Multi-scenario trace names from TransferOptimizer-072026.py's selectable-
# scenario map, e.g. "#3 Switch", "#1 ★OPT Tie 38.8m", "#2 Voltages" — the
# ★OPT tag (only present on the optimal scenario) decodes to a literal '★'
# once the JSON is parsed, so no escape-sequence handling is needed here.
_SCEN_SWITCH_RE = re.compile(r'^#(\d+)\D* Switch$')
_SCEN_TIE_RE = re.compile(r'^#(\d+)\D* Tie ([\d.]+)m$')
_SCEN_VOLT_RE = re.compile(r'^#(\d+)\D* Voltages$')


def _extract_scenario_layers(traces: list) -> list:
    """Per-scenario switch/tie/voltage layers from the new-format (multi-
    scenario selector) Plotly output — one entry per scenario rank found,
    sorted ascending, each carrying its own `optimal` flag. This mirrors
    the "เลือก Scenario" buttons on the Plotly view so the ArcGIS map can
    offer the same picker instead of only ever showing the optimal one.
    """
    switches, ties, volts = {}, {}, {}
    optimal_ranks = set()
    for t in traces:
        name = str(t.get('name') or '')
        m = _SCEN_SWITCH_RE.match(name)
        if m:
            rank = int(m.group(1))
            switches[rank] = t
            if '★' in name:
                optimal_ranks.add(rank)
            continue
        m = _SCEN_TIE_RE.match(name)
        if m:
            rank = int(m.group(1))
            ties[rank] = (t, name)
            if '★' in name:
                optimal_ranks.add(rank)
            continue
        m = _SCEN_VOLT_RE.match(name)
        if m:
            rank = int(m.group(1))
            volts[rank] = t
            if '★' in name:
                optimal_ranks.add(rank)

    scenarios = []
    for rank in sorted(set(switches) | set(ties) | set(volts)):
        entry = {
            'rank': rank, 'optimal': rank in optimal_ranks,
            'switch_opt': None, 'tie': None, 'nodes_voltage': [],
        }
        if rank in switches:
            pts = _trace_pts(switches[rank])
            if len(pts) >= 2:
                entry['switch_opt'] = {'pts': pts, 'label': f'Switch OPEN (#{rank})'}
        if rank in ties:
            t, name = ties[rank]
            pts = _trace_pts(t)
            if len(pts) >= 2:
                entry['tie'] = {'pts': pts, 'label': name}
        if rank in volts:
            entry['nodes_voltage'] = _voltage_nodes_from_trace(volts[rank])
        scenarios.append(entry)
    return scenarios


def parse_plotly_to_map_data(html_path: Path, fac_a: str, fac_b: str) -> dict:
    html = html_path.read_text(encoding='utf-8')
    traces = _extract_plotly_traces(html)

    result = {
        'fac_a': fac_a, 'fac_b': fac_b,
        'tr_a': None, 'tr_b': None,
        'lines_a': [], 'lines_b': [],
        'switch_opt': None, 'tie': None,
        'switches_all': [],
        'nodes_voltage': [],
        'meters_a': [], 'meters_b': [],
        # Per-scenario switch/tie/voltage layers, one entry per feasible
        # scenario — only populated for the new multi-scenario-selector
        # format; stays empty for the old single-scenario format, telling
        # the frontend there's nothing to pick between.
        'scenarios': [],
    }

    tra = _trace_by_name(traces, f'TR_A ({fac_a})')
    if tra and tra.get('x'):
        pt = _ll(tra['x'][0], tra['y'][0])
        if pt:
            result['tr_a'] = {'lon': pt[0], 'lat': pt[1], 'fac': fac_a}

    trb = _trace_by_name(traces, f'TR_B ({fac_b})')
    if trb and trb.get('x'):
        pt = _ll(trb['x'][0], trb['y'][0])
        if pt:
            result['tr_b'] = {'lon': pt[0], 'lat': pt[1], 'fac': fac_b}

    la = _trace_by_name(traces, f'TR_A Lines ({fac_a})')
    if la:
        result['lines_a'] = _lines_to_paths(la.get('x') or [], la.get('y') or [])

    lb = _trace_by_name(traces, f'TR_B Lines ({fac_b})')
    if lb:
        result['lines_b'] = _lines_to_paths(lb.get('x') or [], lb.get('y') or [])

    result['meters_a'] = _meter_points(traces, 'TR_A Meters')
    result['meters_b'] = _meter_points(traces, 'TR_B Meters')

    # ---- Old format (TransferOptimizer.py): single fixed-name traces ----
    sw = _trace_by_name(traces, '[OPT] Switch OPEN')
    if sw is not None:
        pts = _trace_pts(sw)
        if len(pts) >= 2:
            result['switch_opt'] = {'pts': pts, 'label': 'Switch OPEN (optimal)'}

    tie = _trace_by_prefix(traces, 'Tie ')
    if tie is not None:
        pts = _trace_pts(tie)
        if len(pts) >= 2:
            result['tie'] = {'pts': pts, 'label': tie.get('name')}

    result['switches_all'] = [
        pts for t in traces
        if re.match(r'^Switch #\d+$', str(t.get('name') or ''))
        for pts in [_trace_pts(t)] if len(pts) >= 2
    ]

    volt_trace = _trace_by_name(traces, 'Nodes (voltage)')
    if volt_trace is not None:
        result['nodes_voltage'] = _voltage_nodes_from_trace(volt_trace)

    # ---- New format (TransferOptimizer-072026.py): multi-scenario selector ----
    # If nothing above matched, this HTML almost certainly came from the
    # newer optimizer's per-scenario trace naming — try that scheme too
    # instead of leaving the map missing tie/switch/voltage layers.
    if not (result['switch_opt'] or result['tie'] or result['switches_all'] or result['nodes_voltage']):
        scenarios = _extract_scenario_layers(traces)
        result['scenarios'] = scenarios
        result['switches_all'] = [s['switch_opt']['pts'] for s in scenarios if s['switch_opt']]
        default = next((s for s in scenarios if s['optimal']), scenarios[0] if scenarios else None)
        if default:
            result['switch_opt'] = default['switch_opt']
            result['tie'] = default['tie']
            result['nodes_voltage'] = default['nodes_voltage']

    return result


# ------------------------------------------------------------------
# Results table: derive from the xlsx when the JSON export is missing
# ------------------------------------------------------------------

def _build_table_from_xlsx(xlsx_path: Path, fac_a: str, fac_b: str) -> dict | None:
    """Rebuild the per-scenario results table from the xlsx 'Results' sheet.

    Runs from before the JSON export was added only have the xlsx — its
    'Results' sheet carries the same per-scenario columns (see
    save_excel_report in TransferOptimizer.py), so the table can be
    reconstructed from it instead of requiring a full re-run.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if 'Results' not in wb.sheetnames:
            return None
        ws = wb['Results']

        def _num(v):
            return float(v) if v not in (None, '') else None

        scenarios = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if row is None or all(c is None for c in row):
                continue
            cols = list(row) + [None] * 18
            (_rank, status, su, sv, na, nb, tie_dist, kw, cond,
             a_before, b_before, a_after, b_after,
             _a_minv, _b_minv, overall_minv, feasible, error) = cols[:18]

            a_after_n, b_after_n, minv_n = _num(a_after), _num(b_after), _num(overall_minv)
            scenarios.append({
                'rank': i,
                'switch_u': su, 'switch_v': sv,
                'tie_node_a': na, 'tie_node_b': nb,
                'tie_distance_m': round(_num(tie_dist) or 0.0, 1),
                'tie_conductor_size': cond,
                'subtree_kw': round(_num(kw) or 0.0, 1),
                'a_loading_before': round(_num(a_before) or 0.0, 1),
                'a_loading_after': round(a_after_n, 1) if a_after_n is not None else None,
                'b_loading_before': round(_num(b_before) or 0.0, 1),
                'b_loading_after': round(b_after_n, 1) if b_after_n is not None else None,
                'min_v': round(minv_n, 0) if minv_n is not None else None,
                'feasible': (feasible == 'YES'),
                'optimal': (status == '** OPTIMAL **'),
                'error': error or None,
            })
        if not scenarios:
            return None
        return {'fac_a': fac_a, 'fac_b': fac_b, 'scenarios': scenarios}
    finally:
        wb.close()


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def _list_pairs():
    pairs = []
    if not FEASIBLE_DIR.exists():
        return pairs
    for folder in FEASIBLE_DIR.iterdir():
        if not folder.is_dir():
            continue
        m = PAIR_RE.match(folder.name)
        if not m:
            continue
        fac_a, fac_b = m.group(1), m.group(2)
        html_file = folder / f'transfer_{folder.name}.html'
        pairs.append({
            'key': folder.name,
            'fac_a': fac_a,
            'fac_b': fac_b,
            'has_html': html_file.exists(),
            'has_xlsx': (folder / f'transfer_{folder.name}.xlsx').exists(),
            'mtime': datetime.fromtimestamp(folder.stat().st_mtime).strftime('%Y-%m-%d %H:%M'),
        })
    return sorted(pairs, key=lambda x: x['mtime'], reverse=True)


# ------------------------------------------------------------------
# Routes
# ------------------------------------------------------------------

@shareload_bp.route('/shareload')
@login_required
def shareload_list():
    user = session.get('user', {})
    pairs = _list_pairs()
    with _running_lock:
        running = set(_running_jobs)
    return render_template('shareload.html', pairs=pairs, running=running, user=user,
                           regions=sorted(set(Config.REGION_MAPPING.values())),
                           default_region=get_user_region())


@shareload_bp.route('/shareload/run', methods=['POST'])
@login_required
def shareload_run():
    fac_a = request.form.get('fac_a', '').strip()
    fac_b = request.form.get('fac_b', '').strip()
    force = request.form.get('force') == '1'
    region = request.form.get('region', '').strip().upper()

    if not FAC_RE.match(fac_a) or not FAC_RE.match(fac_b):
        return jsonify({'error': 'รูปแบบ FACILITYID ไม่ถูกต้อง (XX-XXXXXX)'}), 400
    if fac_a == fac_b:
        return jsonify({'error': 'FACILITYID A และ B ต้องไม่เป็นตัวเดียวกัน'}), 400
    if region not in set(Config.REGION_MAPPING.values()):
        return jsonify({'error': 'กรุณาเลือกเขต GIS ของหม้อแปลงทั้งคู่'}), 400

    pair_key = f'{fac_a}_{fac_b}'
    out_dir = FEASIBLE_DIR / pair_key
    html_path = out_dir / f'transfer_{pair_key}.html'

    if html_path.exists() and not force:
        return jsonify({'status': 'exists', 'key': pair_key}), 200

    with _running_lock:
        if pair_key in _running_jobs:
            return jsonify({'status': 'running', 'key': pair_key}), 202
        _running_jobs.add(pair_key)

    if force and out_dir.exists():
        # Clear stale output before rerunning — otherwise /shareload/status
        # and /shareload/table keep reporting the previous run's html/table
        # as ready while this run is still in progress or has failed, and
        # the UI never reflects the rerun's real outcome.
        shutil.rmtree(out_dir, ignore_errors=True)

    def _worker():
        try:
            out_dir.mkdir(parents=True, exist_ok=True)
            script = SHARELOAD_DIR / 'run_web.py'
            proc = subprocess.run(
                [sys.executable, str(script), fac_a, fac_b, str(out_dir), region],
                cwd=str(SHARELOAD_DIR),
                capture_output=True, text=True, timeout=600,
                # run_web.py reconfigures its own stdout/stderr to UTF-8; without
                # this, Python decodes the pipes using the OS locale codepage
                # (cp874 on Thai Windows), which crashes on the Thai/Unicode text
                # run_web.py prints, killing the internal reader thread and
                # leaving proc.stderr as None.
                encoding="utf-8", errors="replace",
            )
            if proc.returncode != 0:
                logging.error(f"[shareload] {pair_key} failed:\n{proc.stderr}")
            else:
                logging.info(f"[shareload] {pair_key} done: {proc.stdout.strip()}")
        except Exception:
            logging.exception(f"[shareload] {pair_key} exception")
        finally:
            with _running_lock:
                _running_jobs.discard(pair_key)

    threading.Thread(target=_worker, daemon=True).start()
    return jsonify({'status': 'running', 'key': pair_key}), 202


@shareload_bp.route('/shareload/delete/<pair_key>', methods=['POST'])
@login_required
def shareload_delete(pair_key):
    """Delete a cached run's output folder so it can be regenerated from scratch."""
    if not PAIR_RE.match(pair_key):
        abort(400)
    with _running_lock:
        if pair_key in _running_jobs:
            return jsonify({'error': 'กำลังประมวลผลอยู่ ไม่สามารถลบได้ในขณะนี้'}), 409
    folder = FEASIBLE_DIR / pair_key
    if folder.exists():
        shutil.rmtree(folder, ignore_errors=True)
    return jsonify({'status': 'deleted', 'key': pair_key})


@shareload_bp.route('/shareload/status/<pair_key>')
@login_required
def shareload_status(pair_key):
    if not PAIR_RE.match(pair_key):
        abort(400)
    html_path = FEASIBLE_DIR / pair_key / f'transfer_{pair_key}.html'
    with _running_lock:
        running = pair_key in _running_jobs
    return jsonify({'ready': html_path.exists(), 'running': running})


@shareload_bp.route('/shareload/result/<pair_key>')
@login_required
def shareload_result(pair_key):
    m = PAIR_RE.match(pair_key)
    if not m:
        abort(404)
    fac_a, fac_b = m.group(1), m.group(2)
    folder = FEASIBLE_DIR / pair_key
    user = session.get('user', {})
    return render_template('shareload_result.html',
        pair_key=pair_key, fac_a=fac_a, fac_b=fac_b,
        has_html=(folder / f'transfer_{pair_key}.html').exists(),
        has_xlsx=(folder / f'transfer_{pair_key}.xlsx').exists(),
        has_table=(folder / f'transfer_{pair_key}_results.json').exists(),
        user=user,
    )


@shareload_bp.route('/shareload/table/<pair_key>')
@login_required
def shareload_table(pair_key):
    """Return the per-scenario results table (switch/tie/loading/status) as JSON."""
    m = PAIR_RE.match(pair_key)
    if not m:
        abort(400)
    folder = FEASIBLE_DIR / pair_key
    json_path = folder / f'transfer_{pair_key}_results.json'

    if not json_path.exists():
        # Older runs (from before the JSON export existed) only have the
        # xlsx — rebuild the table from it instead of forcing a re-run.
        xlsx_path = folder / f'transfer_{pair_key}.xlsx'
        if not xlsx_path.exists():
            return jsonify({'scenarios': None}), 200
        try:
            data = _build_table_from_xlsx(xlsx_path, m.group(1), m.group(2))
        except Exception:
            logging.exception(f"[shareload/table] xlsx fallback failed for {pair_key}")
            return jsonify({'scenarios': None}), 200
        if not data:
            return jsonify({'scenarios': None}), 200
        try:
            with open(json_path, 'w', encoding='utf-8') as fh:
                json.dump(data, fh, ensure_ascii=False)
        except Exception:
            logging.exception(f"[shareload/table] failed to cache derived table for {pair_key}")
        return jsonify(data)

    try:
        with open(json_path, encoding='utf-8') as fh:
            data = json.load(fh)
        return jsonify(data)
    except Exception as e:
        logging.exception(f"[shareload/table] read error for {pair_key}")
        return jsonify({'error': str(e)}), 500


@shareload_bp.route('/shareload/data/<pair_key>')
@login_required
def shareload_data(pair_key):
    """Return parsed network data as WGS84 JSON for ArcGIS map rendering."""
    m = PAIR_RE.match(pair_key)
    if not m:
        abort(400)
    fac_a, fac_b = m.group(1), m.group(2)
    html_path = FEASIBLE_DIR / pair_key / f'transfer_{pair_key}.html'
    if not html_path.exists():
        abort(404)
    try:
        data = parse_plotly_to_map_data(html_path, fac_a, fac_b)
        return jsonify(data)
    except Exception as e:
        logging.exception(f"[shareload/data] parse error for {pair_key}")
        return jsonify({'error': str(e)}), 500


@shareload_bp.route('/shareload/view/<pair_key>')
@login_required
def shareload_view_html(pair_key):
    """Serve raw Plotly HTML (fallback/download)."""
    if not PAIR_RE.match(pair_key):
        abort(404)
    html_path = FEASIBLE_DIR / pair_key / f'transfer_{pair_key}.html'
    if not html_path.exists():
        abort(404)
    return send_file(str(html_path), mimetype='text/html')


@shareload_bp.route('/shareload/download/<pair_key>')
@login_required
def shareload_download(pair_key):
    if not PAIR_RE.match(pair_key):
        abort(404)
    xlsx_path = FEASIBLE_DIR / pair_key / f'transfer_{pair_key}.xlsx'
    if not xlsx_path.exists():
        abort(404)
    return send_file(str(xlsx_path), as_attachment=True,
                     download_name=f'transfer_{pair_key}.xlsx')
